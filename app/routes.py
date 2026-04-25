import asyncio
import hmac, hashlib
import os, time, base64, io, json, re, uuid, httpx
from datetime import datetime, timezone

from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File, Form, Request, Header, Depends
from fastapi.responses import StreamingResponse, Response, RedirectResponse
from pydantic import BaseModel
from supabase import create_client, Client

# ── SUPABASE CLIENT ───────────────────────────────────────────────────────────
SUPABASE_URL         = os.getenv("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
SUPABASE_JWT_SECRET  = os.getenv("SUPABASE_JWT_SECRET", "")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY) if SUPABASE_URL else None

# ── JWT MANUAL PARA KLING ────────────────────────────────────────────────────
def _make_kling_jwt(access_key: str, secret_key: str) -> str:
    import base64, json, time
    now = int(time.time())
    header  = base64.urlsafe_b64encode(json.dumps({"alg":"HS256","typ":"JWT"}).encode()).rstrip(b"=").decode()
    payload = base64.urlsafe_b64encode(json.dumps({"iss":access_key,"exp":now+1800,"nbf":now-5}).encode()).rstrip(b"=").decode()
    msg = f"{header}.{payload}".encode()
    sig = base64.urlsafe_b64encode(hmac.new(secret_key.encode(), msg, hashlib.sha256).digest()).rstrip(b"=").decode()
    return f"{header}.{payload}.{sig}"

router = APIRouter()

def parse_expiry(expires_at: str) -> datetime:
    """Parsea plan_expires_at tolerando formatos +00 y +00:00 de Postgres."""
    if not expires_at:
        return datetime.max.replace(tzinfo=timezone.utc)
    s = expires_at.replace("Z", "+00:00")
    s = re.sub(r'\+(\d{2})$', r'+\1:00', s)
    s = re.sub(r'-(\d{2})$', r'-\1:00', s)
    try:
        return datetime.fromisoformat(s)
    except Exception:
        try:
            return datetime.fromisoformat(s.split('+')[0] + '+00:00')
        except Exception:
            return datetime.max.replace(tzinfo=timezone.utc)


# ── API KEYS ─────────────────────────────────────────────────────────────────
GROQ_KEY   = os.getenv("GROQ_API_KEY", "")
TAVILY_KEY = os.getenv("TAVILY_API_KEY", "")
GEMINI_KEY = os.getenv("GEMINI_API_KEY", "")
OPENAI_KEY = os.getenv("OPENAI_API_KEY", "")
APP_URL    = os.getenv("APP_URL", "https://orquesta.up.railway.app")

# ── STRIPE ───────────────────────────────────────────────────────────────────
STRIPE_SECRET_KEY      = os.getenv("STRIPE_SECRET_KEY", "")
STRIPE_WEBHOOK_SECRET  = os.getenv("STRIPE_WEBHOOK_SECRET", "")
STRIPE_PRICE_MONTHLY   = os.getenv("STRIPE_PRICE_MONTHLY", "")
STRIPE_PRICE_ANNUAL    = os.getenv("STRIPE_PRICE_ANNUAL", "")

# ── MERCADOPAGO ───────────────────────────────────────────────────────────────
MP_ACCESS_TOKEN = os.getenv("MERCADOPAGO_ACCESS_TOKEN", "")

# ── LÍMITES ───────────────────────────────────────────────────────────────────
FREE_DAILY_LIMIT = int(os.getenv("FREE_DAILY_LIMIT", "20"))

# ── CRÉDITOS DE VIDEO ─────────────────────────────────────────────────────────
# ── PRECIOS DE VIDEO (FAL.ai → precio a usuario) ─────────────────────────────
# Costo real FAL.ai x 2 = precio al usuario (100% de margen)
VIDEO_PRICES = {
    # (modelo, duracion_s): {"fal_cost": X, "user_price": X*2, "label": "..."}
    ("kling25pro", 5):  {"fal_cost": 0.35, "user_price": 0.70,  "label": "Kling 2.5 Pro · 5s · 1080p"},
    ("kling25pro", 10): {"fal_cost": 0.70, "user_price": 1.40,  "label": "Kling 2.5 Pro · 10s · 1080p"},
    ("kling26pro", 5):  {"fal_cost": 0.70, "user_price": 1.40,  "label": "Kling 2.6 Pro · 5s · Audio nativo"},
    ("kling26pro", 10): {"fal_cost": 1.40, "user_price": 2.80,  "label": "Kling 2.6 Pro · 10s · Audio nativo"},
}
VIDEO_DEFAULT_MODEL   = "kling25pro"
VIDEO_DEFAULT_DURATION = 5

# ── RATE LIMITING por IP ──────────────────────────────────────────────────────
from collections import defaultdict
_ip_requests: dict = defaultdict(list)  # ip -> [timestamps]

def check_rate_limit(ip: str, max_per_minute: int = 20) -> bool:
    """Retorna True si la IP está dentro del límite, False si excedió."""
    now = time.time()
    window = _ip_requests[ip]
    # Limpiar requests de más de 60 segundos
    _ip_requests[ip] = [t for t in window if now - t < 60]
    if len(_ip_requests[ip]) >= max_per_minute:
        return False
    _ip_requests[ip].append(now)
    return True

# ── CACHÉ DE RESPUESTAS ───────────────────────────────────────────────────────
_response_cache: dict = {}  # hash -> (result, timestamp)

def get_cached_response(prompt: str, mode: str) -> str | None:
    """Retorna respuesta cacheada si existe y tiene menos de 1 hora."""
    key = f"{hash(prompt[:100])}_{mode}"
    if key in _response_cache:
        result, ts = _response_cache[key]
        if time.time() - ts < 3600:  # 1 hora
            return result
    return None

def cache_response(prompt: str, mode: str, result: str):
    """Cachea una respuesta general (solo para preguntas genéricas)."""
    key = f"{hash(prompt[:100])}_{mode}"
    _response_cache[key] = (result, time.time())
    # Limpiar caché si tiene más de 500 entradas
    if len(_response_cache) > 500:
        oldest = sorted(_response_cache.keys(), key=lambda k: _response_cache[k][1])[:100]
        for k in oldest:
            del _response_cache[k]

# ── FUNCIONES QUE REQUIEREN PRO ───────────────────────────────────────────────
PRO_ONLY_TASKS = {
    "image_gen", "video_gen", "sound_gen",
    "file_gen_xlsx", "file_gen_docx", "file_gen_pdf",
}

# ─────────────────────────────────────────────────────────────────────────────
#  AUTH HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def verify_jwt(token: str) -> dict:
    """Verifica el JWT de Supabase y retorna el payload."""
    try:
        parts = token.split(".")
        if len(parts) != 3:
            raise ValueError("JWT inválido: estructura incorrecta")

        # Decodificar payload
        pad = parts[1] + "=" * (4 - len(parts[1]) % 4)
        payload = json.loads(base64.urlsafe_b64decode(pad))

        # Verificar expiración
        if "exp" in payload and payload["exp"] < time.time():
            raise ValueError("Token expirado")

        # Verificar firma — Supabase usa HS256
        if SUPABASE_JWT_SECRET:
            signing_input = f"{parts[0]}.{parts[1]}".encode("utf-8")
            # El secret de Supabase puede estar en base64 o como string plano
            secret_str = SUPABASE_JWT_SECRET.strip()
            # Intentar como string UTF-8 directo
            try:
                sig = hmac.new(secret_str.encode("utf-8"), signing_input, hashlib.sha256).digest()
                expected = base64.urlsafe_b64encode(sig).rstrip(b"=").decode()
                if parts[2] == expected:
                    return payload
            except Exception:
                pass
            # Intentar decodificando el secret como base64
            try:
                secret_bytes = base64.b64decode(secret_str + "==")
                sig = hmac.new(secret_bytes, signing_input, hashlib.sha256).digest()
                expected = base64.urlsafe_b64encode(sig).rstrip(b"=").decode()
                if parts[2] == expected:
                    return payload
            except Exception:
                pass
            raise ValueError("Firma JWT inválida")

        return payload
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Token inválido: {str(e)}")

def get_auth_id_from_token(token: str) -> str:
    """Extrae el auth_id (sub) del JWT."""
    payload = verify_jwt(token)
    return payload.get("sub")

async def get_current_user(authorization: str = Header(None)) -> dict:
    """
    Dependency: extrae y valida el usuario desde el header Authorization.
    Retorna los datos del usuario desde public.users.
    """
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="No autenticado")

    token = authorization.replace("Bearer ", "")
    auth_id = get_auth_id_from_token(token)

    if not supabase:
        raise HTTPException(status_code=500, detail="Supabase no configurado")

    result = supabase.table("users").select("*").eq("auth_id", auth_id).single().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    return result.data

async def get_optional_user(authorization: str = Header(None)) -> dict | None:
    """Como get_current_user pero nunca falla — retorna None si hay cualquier error."""
    if not authorization or not authorization.startswith("Bearer "):
        return None
    try:
        token = authorization.replace("Bearer ", "").strip()
        auth_id = get_auth_id_from_token(token)
        if not supabase:
            return None
        result = supabase.table("users").select("*").eq("auth_id", auth_id).single().execute()
        return result.data if result.data else None
    except Exception as e:
        print(f"get_optional_user JWT error: {e}")
        return None

def check_pro_access(user: dict, task: str) -> dict | None:
    """
    Verifica si el usuario puede ejecutar la tarea.
    Retorna None si puede, o un dict con el mensaje de upgrade si no puede.
    Plan Free: chat de texto hasta 20 preguntas.
    Plan Pro: todo ilimitado.
    """
    if not user:
        # Sin login: puede hacer chat de texto general (el frontend controla el límite de 20)
        # Solo bloquear funciones Pro
        if task in PRO_ONLY_TASKS:
            return {
                "blocked": True,
                "reason": "pro_required",
                "message": "Esta funcion requiere Plan Pro. Plan Gratis: solo texto (hasta 20 preguntas). Plan Pro $9/mes: imagenes, videos, archivos, voz y mensajes ilimitados.",
                "cta": "Activar Pro",
                "cta_url": "/pricing"
            }
        # Chat de texto sin login: permitido
        return None

    is_pro = (
        user.get("plan") == "pro" and (
            user.get("plan_expires_at") is None or
            parse_expiry(user["plan_expires_at"]) > datetime.now(timezone.utc)
        )
    )

    if task in PRO_ONLY_TASKS and not is_pro:
        task_names = {
            "image_gen":      "generación de imágenes",
            "video_gen":      "generación de videos",
            "sound_gen":      "generación de audio",
            "file_gen_xlsx":  "creación de archivos Excel",
            "file_gen_docx":  "creación de archivos Word",
            "file_gen_pdf":   "creación de archivos PDF",
            "realtime":       "búsqueda web en tiempo real",
        }
        feature = task_names.get(task, "esta función")
        return {
            "blocked": True,
            "reason": "pro_required",
            "message": (
                f"⚡ La {feature} es exclusiva del plan **Pro**.\n\n"
                f"Con Pro tenés acceso ilimitado a imágenes, videos, archivos, "
                f"búsqueda web, voz y mucho más.\n\n"
                f"**Plan Pro**: $9 USD/mes · $95 USD/año *(ahorrás $13)*"
            ),
            "cta": "Activar Pro ahora",
            "cta_url": "/pricing"
        }

    # Verificar límite diario para usuarios Free
    if not is_pro:
        daily_count = user.get("daily_message_count", 0)

        if daily_count >= FREE_DAILY_LIMIT:
            return {
                "blocked": True,
                "reason": "daily_limit",
                "message": (
                    f"📊 Alcanzaste el límite de **{FREE_DAILY_LIMIT} mensajes diarios** del plan Free.\n\n"
                    f"Con **Pro** tenés mensajes ilimitados, más imágenes, videos, archivos y voz.\n\n"
                    f"**Plan Pro**: $9 USD/mes · $95 USD/año *(ahorrás $13)*"
                ),
                "cta": "Activar Pro — $9/mes",
                "cta_url": "/pricing"
            }

    return None  # Puede enviar


# ─────────────────────────────────────────────────────────────────────────────
#  SYSTEM PROMPT
# ─────────────────────────────────────────────────────────────────────────────

BASE_SYSTEM = """Sos Orquesta. Una IA diseñada por Horacio Basly (HB Soluciones Informáticas).

PERSONALIDAD:
- Hablás como una persona real, cálida, directa y con criterio propio
- Usás el nombre del usuario cuando es natural, no en cada frase
- Recordás TODO lo que se habló en la conversación actual y lo usás
- Sos honesta: si algo falló o no pudiste generarlo, lo decís claramente
- NUNCA repetís tu nombre salvo que te lo pregunten directamente
- NUNCA decís frases vacías como "¡Claro!", "¡Por supuesto!", "¡Entendido!"
- NUNCA terminás con "¿En qué más puedo ayudarte?" ni similares

CAPACIDADES REALES (sé honesta sobre estas):
- Texto e información: siempre disponible, sin límites
- Imágenes IA: Pollinations Flux (gratis, siempre activo)
- Videos IA: Replicate/Minimax cuando hay créditos disponibles. Si no hay créditos, lo decís claramente y ofrecés la descripción
- Archivos Excel/Word/PDF: siempre disponible
- Búsqueda web: disponible con Tavily
- Voz TTS: disponible con gTTS (gratis)

COHERENCIA CONVERSACIONAL (CRÍTICO):
- Leé TODO el historial antes de responder
- Si el usuario dice "dame el archivo", "el mp4", "el video", "eso que me dijiste" → es continuación directa del tema anterior
- Si prometiste generar algo y el sistema lo intentó, informá el resultado real: si salió bien, confirmá; si falló, explicá por qué
- NUNCA volvás a pedir información que el usuario ya te dio en la misma conversación
- Si el usuario repite un pedido es porque no lo recibió — no le pidas más datos, intentá de nuevo o explicá el problema técnico real

HONESTIDAD SOBRE VIDEOS:
- Cuando generás un video exitosamente → mostrás el reproductor con el video real
- Cuando falla por falta de créditos → decís: "Intenté generarlo pero las APIs de video no tienen créditos disponibles ahora. Puedo darte una descripción detallada o esperar a que se recarguen."
- NUNCA digas que "estás generando" si el sistema no tiene APIs activas

SISTEMA DE AUTO-MEJORAMIENTO:
- Tenés un sistema activo que monitorea errores y propone mejoras a tu dueño Horacio
- Cuando el usuario te reporta un error, lo registrás automáticamente y disparás un análisis
- El análisis genera una propuesta que se envía por email a Horacio (ms.horasoft@gmail.com)
- IMPORTANTE: NO digas "te acabo de enviar un email" — el email lo envía el sistema en background, no vos directamente. Decí en cambio: "Registré el error y el sistema de auto-mejoramiento va a enviarle una propuesta a Horacio para revisión."
- Horacio aprueba o rechaza con un clic — si aprueba, el cambio se aplica automáticamente en GitHub

ESTILO DE RESPUESTA:
- Respuestas concisas cuando la pregunta es simple
- Respuestas detalladas cuando el tema lo requiere
- Nunca usés markdown excesivo — solo cuando realmente ayuda a la lectura
- En español argentino informal cuando el usuario habla así
- Resumen ejecutivo al inicio solo en respuestas muy largas"""

MODE_PROMPTS = {
    "tecnico":  "\n\nMODO TÉCNICO: Sos el mejor ingeniero/científico del mundo en el área. Incluí valores exactos, fórmulas con variables definidas, normas específicas (ISO/ASTM/IEC/API), rangos de tolerancia, casos de fallo y árbol de causas.",
    "creativo": "\n\nMODO CREATIVO: Sos director creativo senior de nivel mundial. Original, inesperado, memorable. Múltiples variantes con ejemplos concretos. Adaptá el tono exactamente al contexto.",
    "codigo":   "\n\nMODO CÓDIGO: Sos el mejor developer del mundo (Google+Meta+Netflix nivel). Código production-ready, typed, con manejo exhaustivo de errores, tests sugeridos. Explicás cada decisión de diseño. Señalás anti-patterns.",
}

def get_system(mode, username="", history=None):
    sys = BASE_SYSTEM
    if username:
        sys += f"\n\nEl usuario se llama {username}."
    # Agregar resumen del contexto de conversación si hay historial
    if history and len(history) > 2:
        recent_topics = []
        for m in history[-6:]:
            content = m.get("content","")[:80]
            role = "Usuario" if m.get("role") == "user" else "Orquesta"
            recent_topics.append(f"{role}: {content}")
        if recent_topics:
            sys += f"\n\nContexto reciente de la conversación:\n" + "\n".join(recent_topics)
            sys += "\n\nUsá este contexto para dar respuestas coherentes y no repetir información ya dada."
    sys += MODE_PROMPTS.get(mode, "")
    return sys


# ─────────────────────────────────────────────────────────────────────────────
#  CLASIFICADORES
# ─────────────────────────────────────────────────────────────────────────────

VIDEO_GEN_KW = [
    # Frases directas
    "genera un video","generá un video","crea un video","creá un video",
    "hace un video","hacé un video","haceme un video","generame un video",
    "quiero un video","necesito un video","dame un video",
    "generate video","create video","make a video","make me a video",
    # Con palabras intermedias
    "me generes un video","me hagas un video","me crees un video",
    "me generés un video","generar un video","crear un video",
    "video animado","video corto","video publicitario","video promocional",
    "video en 4k","video 4k","video hd","video uhd","video de",
    # Otros formatos
    "animación de","animación con","animate","clip de","short video",
    "film","película corta","animá","animar","renderizá un video","render video",
    "dame el video","el video de","video con",
]
IMAGE_GEN_KW = [
    "genera una imagen","generá una imagen","crea una imagen","creá una imagen",
    "dibuja","dibujá","ilustra","ilustrá","imagen de","foto de","fotografía de",
    "generate image","create image","draw","make an image","picture of","render",
    "diseña","diseñá","hazme una imagen","create a photo","make a photo",
    "quiero una imagen","quiero ver","mostrame","mostrarme","visualizá","visualiza",
    "haceme una imagen","hacé una imagen","necesito una imagen","generame","generame una",
    "foto de","pintura de","retrato de","ilustración de","arte de","artwork",
    "imagina","imaginá","una foto","una imagen","un dibujo","un retrato",
]
IMAGE_EDIT_KW = [
    "modifica","modificá","cambia","cambiá","reemplaza","reemplazá","edita","editá",
    "borra","añade","quita","quitá","pon","pone","transforma","edit","change",
    "replace","remove","swap","put","add","convert","transform",
]
REALTIME_KW = [
    "hoy","ahora","actual","actualmente","últimas","ultimo","última","esta semana",
    "esta noche","ayer","mañana","reciente","trending","today","now","latest","current",
    "yesterday","tonight","this week","partido","resultado","formación","alineación",
    "ganó","perdió","empató","score","gol","goles","fixture","tabla","clasificación",
    "champions","copa","mundial","liga","torneo","jugó","juega","derrota","victoria",
    "precio","cotización","dólar","euro","peso","bitcoin","crypto","cuánto cuesta",
    "cuánto vale","cotizan","bolsa","acciones","clima","temperatura","pronóstico",
    "lluvia","weather","forecast","noticias","noticia","news","murió","nació","lanzó",
    "salió","eligieron","anunció","declaró","quién ganó","quién es el",
]
CODE_KW = [
    "código","code","función","function","script","python","javascript","typescript",
    "bug","debug","clase","class","algoritmo","sql","html","css","api","json","regex",
    "bash","programa","programar","endpoint","database","query","loop","array","error en",
]
ANALYSIS_KW = [
    "analiza","compare","compara","evalúa","pros","contras","diferencia entre",
    "ventajas","desventajas","estrategia","qué opinas","qué pensás","cuál es mejor",
    "recomienda","conviene","debería","vale la pena","cómo se compara",
]
TRANSLATE_KW = [
    "traduce","traducí","translate","traducción al","how do you say","cómo se dice",
]
FILE_GEN_KW = {
    "xlsx": ["excel","planilla","spreadsheet","hoja de calculo","hoja de cálculo",".xlsx",
             "tabla excel","presupuesto excel","reporte excel","plantilla excel","generar excel",
             "crea excel","haceme un excel","haceme una planilla","plantilla de excel",
             "crea una planilla","generá una planilla","creá un excel"],
    "docx": ["word",".docx","documento word","informe word","cv en word","curriculum en word",
             "carta word","contrato word","memo word","plantilla word","generar word",
             "crea un word","haceme un word","documento de texto","redacta un documento",
             "crea un documento"],
    "pdf":  ["pdf",".pdf","en pdf","como pdf","generar pdf","crea un pdf","haceme un pdf",
             "informe pdf","reporte pdf","cv pdf","curriculum pdf","documento pdf",
             "carta pdf","reporte en pdf"],
}
CV_KW = ["curriculum","currículum","cv ","resume","hoja de vida"]
SOUND_KW = [
    # Pedidos explícitos de música/canción — muy específicos primero
    "generá una canción","genera una cancion","genera una canción","crea una canción",
    "componé una canción","componé una cancion","haceme una canción","haceme una cancion",
    "generame una cancion","generame una canción","generame un tema",
    "canción sobre","cancion sobre","song about","make a song","create a song",
    "generate music","generate a song","música de","musica de","tema musical",
    "compone una cancion","compone una canción","componer una canción",
    # Géneros musicales con verbo de generación (detección combinada abajo)
    "una cumbia","una balada","un reggaeton","un rock","un jazz","un rap","un trap",
    "un folklore","una chacarera","un tango","un vals","una milonga",
    # Audio genérico
    "genera un sonido","generá un sonido","crea un sonido",
    "genera audio","generá audio","crea audio","sound effect",
    "efecto de sonido","beat","melodía","melodia","generate sound","create sound",
]

# Géneros solos — solo cuentan si van con un verbo de generación
MUSIC_GENRES = ["cumbia","reggaeton","salsa","bachata","folklore","chacarera","tango",
                "balada","bolero","trap","rap","jazz","blues","rock","pop","metal"]

def is_music_request(prompt: str) -> bool:
    p = prompt.lower()
    if any(k in p for k in SOUND_KW):
        return True
    # Género + verbo de generación
    gen_verbs = ["genera","generá","crea","creá","hacé","haceme","generame",
                 "componé","compone","quiero","dame","fagas","hagas"]
    has_verb = any(v in p for v in gen_verbs)
    has_genre = any(g in p for g in MUSIC_GENRES)
    # También "canción" o "tema" o "song" + verbo
    has_song_word = any(w in p for w in ["canción","cancion","tema","song","música","musica"])
    return has_verb and (has_genre or has_song_word)


def detect_file_type(prompt):
    p = prompt.lower()
    if any(k in p for k in CV_KW):
        if any(k in p for k in FILE_GEN_KW["docx"]): return "docx"
        if any(k in p for k in FILE_GEN_KW["xlsx"]): return "xlsx"
        return "pdf"
    for ftype, kws in FILE_GEN_KW.items():
        if any(k in p for k in kws): return ftype
    return None

def classify(prompt, mode, history=None):
    if mode == "codigo":   return "code"
    if mode == "creativo": return "creative"
    if mode == "tecnico":  return "technical"
    p = prompt.lower()

    # DETECCIÓN TEMPRANA: quejas, feedback, errores detectados → siempre general
    feedback_kw = [
        "detecté", "detecte", "encontré", "encontre", "noté", "note",
        "hay un error", "tiene un error", "está fallando", "esta fallando",
        "no funciona", "no está funcionando", "problema con", "falla en",
        "necesito que analices", "analiza estos errores", "propone una mejora",
        "no son premium", "son básicos", "son basicos", "son malos",
        "mala calidad", "no me gusta", "mejorar esto", "esto está mal",
        "esto esta mal", "necesito que mejores", "autorepara", "auto-repara",
        "autorepar", "funcionamiento", "errores en tu", "errores en mi",
        # Frases de reporte explícito
        "reporta esta", "reportá esta", "reportar esta", "reportes esta",
        "reporta el", "reportá el", "reportar el", "reportes el",
        "reporta la", "reportá la", "reportar la", "reportes la",
        "reporta esto", "reportá esto",
        "a horacio", "para horacio", "avisá a", "avisa a",
        "automejoramiento", "auto mejoramiento", "sistema de mejora",
        "sistema de auto", "self-improv", "self improv",
        "esta falla", "este fallo", "este error", "esta falla",
        "información detallada", "sin limites", "sin límites",
        "todo lo que averigues", "que averigues",
    ]
    if any(k in p for k in feedback_kw):
        return "general"  # Feedback/queja siempre va a chat general

    if history:
        recent = [m.get("content","").lower() for m in history[-4:]]
        recent_joined = " ".join(recent)

        # Si el contexto reciente tiene código/python/archivo, priorizar eso
        code_context = ["python","código","código","archivo","routes","script",".py",".js","función","class","def ","import "]
        if any(k in recent_joined for k in code_context):
            file_req = ["entregame","el archivo","el código","generame","el .py","el script","dame el","pasame"]
            if any(k in p for k in file_req):
                ftype = detect_file_type(p)
                if ftype: return f"file_gen_{ftype}"
                return "code"

        # Video followup — solo si el contexto reciente tiene video Y el pedido es claro
        video_followup = ["el mp4","el video","reproducir","ver el video","el link del video","la url del video"]
        if any(k in p for k in video_followup):
            if any(k in recent_joined for k in ["video","mp4","kling","minimax","luma","replicate","generando"]):
                return "video_gen"
        # "dame el archivo" solo es video si el contexto es explícitamente de video
        if p.strip() in ["dame el archivo","el archivo","dámelo","entregamelo","mandamelo"]:
            if any(k in recent_joined for k in ["video","mp4","kling","minimax"]):
                return "video_gen"
            elif any(k in recent_joined for k in ["excel","word","pdf","planilla","documento"]):
                ftype = "xlsx" if "excel" in recent_joined or "planilla" in recent_joined else                         "docx" if "word" in recent_joined else "pdf"
                return f"file_gen_{ftype}"

        img_followup = ["la imagen","la foto","no carga","no se ve","otro estilo","más oscura","más grande"]
        if any(k in p for k in img_followup):
            if any(k in recent_joined for k in ["imagen","foto","dall-e","pollinations","generada"]):
                return "image_gen"

    ftype = detect_file_type(p)
    if ftype: return f"file_gen_{ftype}"

    has_video_intent = any(k in p for k in VIDEO_GEN_KW)
    # "video" solo + verbo de generación explícito también cuenta
    gen_verbs = ["genera","generá","crea","creá","hacé","hace","quiero","necesito","dame","haceme","generame","fagas","hagas","generes"]
    if not has_video_intent and "video" in p and any(v in p for v in gen_verbs):
        has_video_intent = True

    if has_video_intent: return "video_gen"
    if is_music_request(p): return "sound_gen"   # música ANTES que imagen
    if any(k in p for k in IMAGE_GEN_KW): return "image_gen"
    if any(k in p for k in TRANSLATE_KW): return "translate"
    if " vs " in p or " contra " in p: return "realtime"
    if any(x in p for x in ["cómo salió","como salio","cómo quedó","qué pasó","que paso"]): return "realtime"
    if any(k in p for k in REALTIME_KW): return "realtime"
    if any(k in p for k in CODE_KW): return "code"
    if any(k in p for k in ANALYSIS_KW): return "analysis"
    return "general"

TASK_LABELS = {
    "image_gen":"openai · dall-e-3","realtime":"orquesta · multi-search","code":"groq · llama 3.3",
    "technical":"groq · mixtral","analysis":"orquesta · análisis+web","creative":"groq · llama 3.3",
    "translate":"groq · traductor","general":"groq · llama 3.3","sound_gen":"orquesta · audio",
    "file_gen_xlsx":"orquesta · excel","file_gen_docx":"orquesta · word","file_gen_pdf":"orquesta · pdf",
}
TASK_MODELS = {
    "code":"llama-3.3-70b-versatile","technical":"llama-3.1-8b-instant","analysis":"llama-3.1-8b-instant",
    "creative":"llama-3.3-70b-versatile","translate":"llama-3.3-70b-versatile",
    "general":"llama-3.3-70b-versatile","realtime":"llama-3.3-70b-versatile",
    "file_gen_xlsx":"llama-3.3-70b-versatile","file_gen_docx":"llama-3.3-70b-versatile",
    "file_gen_pdf":"llama-3.3-70b-versatile",
}

# ─────────────────────────────────────────────────────────────────────────────
#  API CALLERS
# ─────────────────────────────────────────────────────────────────────────────

async def call_groq(messages, model="llama-3.3-70b-versatile"):
    async with httpx.AsyncClient(timeout=60) as c:
        r = await c.post("https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {GROQ_KEY}"},
            json={"model": model, "messages": messages, "max_tokens": 8192, "temperature": 0.7})
        d = r.json()
        if not r.is_success: raise Exception(d.get("error",{}).get("message","Groq error"))
        return d["choices"][0]["message"]["content"]

async def call_gemini(prompt):
    models = ["gemini-2.5-flash"]
    for model in models:
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={GEMINI_KEY}"
            async with httpx.AsyncClient(timeout=35) as c:
                r = await c.post(url, json={"contents":[{"parts":[{"text":prompt}]}],"generationConfig":{"maxOutputTokens":8192}})
                d = r.json()
                if r.is_success:
                    return d["candidates"][0]["content"]["parts"][0]["text"]
                if "UNAVAILABLE" in str(d) or "503" in str(d):
                    continue
                raise Exception(str(d))
        except Exception as e:
            if "UNAVAILABLE" in str(e) or "503" in str(e):
                continue
            raise
    raise Exception("Todos los modelos Gemini no disponibles")

async def call_gemini_vision(prompt, b64, mime):
    # Intentar múltiples modelos Gemini con fallback
    models = ["gemini-2.5-flash"]
    last_error = None
    for model in models:
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={GEMINI_KEY}"
            payload = {"contents":[{"parts":[{"inline_data":{"mime_type":mime,"data":b64}},{"text":prompt}]}],"generationConfig":{"maxOutputTokens":4096,"temperature":0.3}}
            async with httpx.AsyncClient(timeout=45) as c:
                r = await c.post(url, json=payload)
                d = r.json()
                if r.is_success:
                    return d["candidates"][0]["content"]["parts"][0]["text"]
                err = str(d)
                if "UNAVAILABLE" in err or "503" in err or "overloaded" in err.lower() or "high demand" in err.lower():
                    last_error = err
                    await asyncio.sleep(3)  # Esperar antes de probar siguiente
                    continue
                raise Exception(err)
        except Exception as e:
            last_error = str(e)
            if "UNAVAILABLE" in str(e) or "503" in str(e) or "high demand" in str(e).lower():
                await asyncio.sleep(3)
                continue
            raise
    # Fallback 1: GPT-4o Vision (si tiene créditos)
    if OPENAI_KEY:
        try:
            msgs = [
                {"role":"system","content":"Analyze the image/document and respond helpfully in Spanish."},
                {"role":"user","content":[
                    {"type":"image_url","image_url":{"url":f"data:{mime};base64,{b64}"}},
                    {"type":"text","text":prompt}
                ]}
            ]
            async with httpx.AsyncClient(timeout=45) as c:
                r = await c.post("https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {OPENAI_KEY}"},
                    json={"model":"gpt-4o","messages":msgs,"max_tokens":2000})
                d = r.json()
                if r.is_success:
                    return d["choices"][0]["message"]["content"]
        except Exception as e:
            print(f"GPT-4o Vision error: {e}")

    # Fallback 2: Groq LLaVA (gratis, sin límite)
    if GROQ_KEY:
        try:
            async with httpx.AsyncClient(timeout=45) as c:
                r = await c.post(
                    "https://api.groq.com/openai/v1/chat/completions",
                    headers={"Authorization": f"Bearer {GROQ_KEY}"},
                    json={
                        "model": "meta-llama/llama-4-scout-17b-16e-instruct",
                        "messages": [{
                            "role": "user",
                            "content": [
                                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                                {"type": "text", "text": prompt}
                            ]
                        }],
                        "max_tokens": 2000
                    }
                )
                d = r.json()
                if r.is_success:
                    print("✅ Groq Vision OK")
                    return d["choices"][0]["message"]["content"]
                print(f"Groq Vision: {r.status_code} {str(d)[:100]}")
        except Exception as e:
            print(f"Groq Vision error: {e}")

    raise Exception(f"Servicio de análisis de imágenes temporalmente no disponible. Intentá en unos minutos.")

async def call_tavily(query, depth="advanced", max_results=8):
    async with httpx.AsyncClient(timeout=20) as c:
        r = await c.post("https://api.tavily.com/search",
            json={"api_key":TAVILY_KEY,"query":query,"search_depth":depth,"max_results":max_results,"include_answer":True})
        d = r.json()
        if not r.is_success: raise Exception(d.get("message","Tavily error"))
        answer = d.get("answer","")
        results = d.get("results",[])
        ctx = f"Respuesta directa: {answer}\n\n" if answer else ""
        ctx += "\n\n".join(f"[{r['title']}]\n{r['content'][:600]}" for r in results[:6])
        return ctx

async def call_openai_image_gen(prompt):
    async with httpx.AsyncClient(timeout=60) as c:
        r = await c.post("https://api.openai.com/v1/images/generations",
            headers={"Authorization": f"Bearer {OPENAI_KEY}"},
            json={"model":"dall-e-3","prompt":prompt,"n":1,"size":"1024x1024","quality":"standard"})
        d = r.json()
        if not r.is_success: raise Exception(d.get("error",{}).get("message","OpenAI error"))
        return d["data"][0]["url"]

async def call_openai_image_edit(image_bytes, prompt):
    import urllib.parse
    b64 = base64.b64encode(image_bytes).decode()

    # 1. Intentar con OpenAI GPT-4o + DALL-E (requiere créditos)
    if OPENAI_KEY:
        try:
            msgs = [
                {"role":"system","content":"Analyze the image and create a detailed prompt recreating it WITH the modification. Reply ONLY with the prompt in English."},
                {"role":"user","content":[{"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{b64}"}},{"type":"text","text":f"Modification: {prompt}"}]}
            ]
            async with httpx.AsyncClient(timeout=40) as c:
                r = await c.post("https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {OPENAI_KEY}"},
                    json={"model":"gpt-4o","messages":msgs,"max_tokens":600})
                d = r.json()
                if r.is_success:
                    new_prompt = d["choices"][0]["message"]["content"]
                    return await call_openai_image_gen(new_prompt)
        except:
            pass

    # 2. Usar Gemini para analizar la imagen y Pollinations para generar
    if GEMINI_KEY:
        try:
            analysis_prompt = f"Analyze this image in detail and create a prompt to recreate it WITH this modification: {prompt}. Reply ONLY with the English prompt, max 200 words."
            new_prompt = await call_gemini_vision(analysis_prompt, b64, "image/jpeg")
            seed = int(time.time()) % 99999
            encoded = urllib.parse.quote(new_prompt[:400])
            url = f"https://image.pollinations.ai/prompt/{encoded}?width=1024&height=1024&nologo=true&enhance=true&seed={seed}&model=flux&nofeed=true"
            return url
        except Exception as e:
            raise Exception(f"Error al editar imagen: {str(e)[:100]}")

    raise Exception("Para editar imágenes configurá GEMINI_API_KEY o cargá créditos en OpenAI")

async def call_openai_tts(text, voice="nova"):
    clean = text[:4000].strip()

    # 1. OpenAI TTS (mejor calidad, requiere créditos)
    if OPENAI_KEY:
        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.post("https://api.openai.com/v1/audio/speech",
                    headers={"Authorization": f"Bearer {OPENAI_KEY}"},
                    json={"model":"tts-1","input":clean,"voice":voice,"response_format":"mp3"})
                if r.is_success:
                    print("TTS: OpenAI OK")
                    return r.content
                print(f"OpenAI TTS: {r.status_code} {r.text[:100]}")
        except Exception as e:
            print(f"OpenAI TTS error: {e}")

    # 2. ElevenLabs (tier gratis disponible en elevenlabs.io)
    ELEVENLABS_KEY = os.getenv("ELEVENLABS_API_KEY", "")
    if ELEVENLABS_KEY:
        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.post(
                    "https://api.elevenlabs.io/v1/text-to-speech/21m00Tcm4TlvDq8ikWAM",
                    headers={"xi-api-key": ELEVENLABS_KEY, "Content-Type": "application/json"},
                    json={"text": clean[:2500], "model_id": "eleven_multilingual_v2",
                          "voice_settings": {"stability": 0.5, "similarity_boost": 0.75}}
                )
                if r.is_success:
                    print("TTS: ElevenLabs OK")
                    return r.content
        except Exception as e:
            print(f"ElevenLabs TTS error: {e}")

    # 3. gTTS - Google Text to Speech (100% gratis, sin API key)
    try:
        from gtts import gTTS
        import io as _io
        # gTTS soporta textos largos — usar hasta 3000 chars
        tts_text = clean[:3000]
        tts = gTTS(text=tts_text, lang='es', slow=False)
        buf = _io.BytesIO()
        tts.write_to_fp(buf)
        buf.seek(0)
        audio_bytes = buf.read()
        if len(audio_bytes) > 1000:
            print(f"TTS: gTTS OK ({len(tts_text)} chars, {len(audio_bytes)} bytes)")
            return audio_bytes
    except Exception as e:
        print(f"gTTS error: {e}")

    # 4. Groq TTS (si tienen endpoint)
    if GROQ_KEY:
        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.post(
                    "https://api.groq.com/openai/v1/audio/speech",
                    headers={"Authorization": f"Bearer {GROQ_KEY}"},
                    json={"model": "playai-tts", "input": clean[:1000], "voice": "Celeste-PlayAI", "response_format": "mp3"}
                )
                if r.is_success and len(r.content) > 1000:
                    print("TTS: Groq OK")
                    return r.content
                print(f"Groq TTS: {r.status_code} {r.text[:100]}")
        except Exception as e:
            print(f"Groq TTS error: {e}")

    raise Exception("No hay servicio TTS disponible — configurá ELEVENLABS_API_KEY o cargá créditos en OpenAI")

async def call_openai_stt(audio_bytes, filename):
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post("https://api.openai.com/v1/audio/transcriptions",
            headers={"Authorization": f"Bearer {OPENAI_KEY}"},
            files={"file":(filename, audio_bytes, "audio/webm")}, data={"model":"whisper-1"})
        d = r.json()
        if not r.is_success: raise Exception(d.get("error",{}).get("message","STT error"))
        return d["text"]

def build_messages(system, history, prompt):
    msgs = [{"role":"system","content":system}]
    # Incluir historial completo — hasta 20 mensajes para mantener contexto
    valid_history = [m for m in history if m.get("role") in ("user","assistant") and m.get("content")]
    recent = valid_history[-20:]
    for m in recent:
        msgs.append({"role": m["role"], "content": m["content"]})
    msgs.append({"role":"user","content":prompt})
    return msgs

async def groq_with_fallback(messages, model, use_gemini_fallback=True):
    try:
        return await call_groq(messages, model), model
    except Exception:
        try:
            # Usar modelo alternativo con mayor límite de tokens
            alt = "llama-3.3-70b-versatile" if model != "llama-3.3-70b-versatile" else "llama-3.1-8b-instant"
            return await call_groq(messages, alt), alt
        except Exception:
            if GEMINI_KEY and use_gemini_fallback:
                sys_c = next((m["content"] for m in messages if m["role"]=="system"), "")
                usr_c = next((m["content"] for m in reversed(messages) if m["role"]=="user"), "")
                result = await call_gemini(f"{sys_c}\n\n{usr_c}")
                return result, "gemini"
            raise

# ─────────────────────────────────────────────────────────────────────────────
#  IMAGE GENERATION
# ─────────────────────────────────────────────────────────────────────────────

async def generate_image_smart(prompt: str) -> tuple[str, str, str]:
    import urllib.parse

    async def enhance(p):
        try:
            msgs = [
                {"role":"system","content":"You are an expert image prompt engineer. Rewrite the user's request as a vivid, detailed DALL-E/Flux prompt in English. Add: artistic style, lighting, composition, quality descriptors (photorealistic, 8K, cinematic, detailed). Max 200 words. Reply ONLY with the enhanced prompt."},
                {"role":"user","content":f"Request: {p}"}
            ]
            enhanced, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")
            return enhanced.strip()[:400]
        except:
            return p

    enhanced = await enhance(prompt)
    seed = int(time.time()) % 99999

    # 1. Pollinations Flux (gratuito, sin API key) — primera opción
    try:
        encoded = urllib.parse.quote(enhanced)
        url = f"https://image.pollinations.ai/prompt/{encoded}?width=1024&height=1024&nologo=true&enhance=true&seed={seed}&model=flux&nofeed=true"
        return url, "🎨 Imagen generada con **Pollinations AI** (Flux).", "pollinations · flux"
    except:
        pass

    # 2. Pollinations SDXL
    try:
        encoded = urllib.parse.quote(enhanced)
        url = f"https://image.pollinations.ai/prompt/{encoded}?width=1024&height=1024&nologo=true&seed={seed}&model=turbo&nofeed=true"
        return url, "🎨 Imagen generada con **Pollinations AI** (Turbo).", "pollinations · turbo"
    except:
        pass

    # 3. OpenAI DALL-E 3 (requiere créditos)
    if OPENAI_KEY:
        try:
            url = await call_openai_image_gen(enhanced)
            return url, "✨ Imagen generada con **DALL-E 3**.", "openai · dall-e-3"
        except:
            pass

    # 4. Fallback básico
    encoded = urllib.parse.quote(prompt[:300])
    url = f"https://image.pollinations.ai/prompt/{encoded}?width=1024&height=1024&nologo=true&seed={seed}&nofeed=true"
    return url, "🎨 Imagen generada.", "pollinations · imagen"


# ─────────────────────────────────────────────────────────────────────────────
#  CRÉDITOS DE VIDEO
# ─────────────────────────────────────────────────────────────────────────────

async def get_video_credits(user: dict) -> dict:
    """
    Retorna el saldo de video del usuario en USD desde tabla video_balance.
    Estructura: {"balance_usd": float, "is_pro": bool}
    """
    if not user or not supabase:
        return {"balance_usd": 0.0, "is_pro": False}

    is_pro = (
        user.get("plan") == "pro" and (
            user.get("plan_expires_at") is None or
            parse_expiry(user["plan_expires_at"]) > datetime.now(timezone.utc)
        )
    )

    try:
        # maybeSingle() no lanza error si no hay filas (a diferencia de single())
        result = supabase.table("video_balance")             .select("balance_usd")             .eq("user_id", user["id"])             .maybe_single()             .execute()
        balance = round(float(result.data.get("balance_usd", 0.0)), 4) if result.data else 0.0
        print(f"✅ video_balance: ${balance:.4f} USD")
        return {"balance_usd": balance, "is_pro": is_pro}
    except Exception as e:
        print(f"video_balance lookup error: {e}")
        # Si falla, devolver 0 sin tirar excepción
        return {"balance_usd": 0.0, "is_pro": is_pro}


async def consume_video_credit(user: dict, model_key: str, duration_s: int, prompt_preview: str) -> bool:
    """
    Descuenta el costo del video en USD del saldo del usuario.
    Registra el uso en video_usage para historial.
    """
    if not user or not supabase:
        return True

    price_info = VIDEO_PRICES.get((model_key, duration_s), VIDEO_PRICES[("kling25pro", 5)])
    cost = price_info["fal_cost"]  # costo real que pagamos a FAL
    user_price = price_info["user_price"]  # lo que pagó el usuario

    try:
        # Descontar del saldo — usar RPC para atomicidad
        supabase.rpc("deduct_video_balance", {
            "p_user_id": user["id"],
            "p_amount": user_price
        }).execute()
        print(f"✅ Video balance deducido: ${user_price:.2f} de {user['id'][:8]}")
    except Exception as e:
        print(f"deduct_video_balance error: {e}")
        # Intentar update directo como fallback
        try:
            current = supabase.table("video_balance").select("balance_usd").eq("user_id", user["id"]).single().execute()
            new_bal = round(max(0, float(current.data.get("balance_usd", 0)) - user_price), 4)
            supabase.table("video_balance").update({"balance_usd": new_bal}).eq("user_id", user["id"]).execute()
        except Exception as e2:
            print(f"video_balance update fallback error: {e2}")

    # Registrar en historial
    try:
        supabase.table("video_usage").insert({
            "user_id": user["id"],
            "model_key": model_key,
            "duration_s": duration_s,
            "fal_cost_usd": cost,
            "user_price_usd": user_price,
            "prompt_preview": prompt_preview[:100],
            "created_at": datetime.now(timezone.utc).isoformat(),
        }).execute()
    except Exception as e:
        print(f"video_usage insert error (ok): {e}")

    return True


# ─────────────────────────────────────────────────────────────────────────────
#  VIDEO GENERATION  — cascada: Motor propio → Replicate → FAL.ai → ModelsLab → descripción
# ─────────────────────────────────────────────────────────────────────────────

async def generate_video_smart(
    prompt: str,
    history=None,
    mode: str = "general",
    username: str = "",
    user: dict = None,
    model_key: str = "kling25pro",
    duration_s: int = 5,
) -> tuple[str, str, str]:
    """
    Genera video usando FAL.ai — Kling 2.5 Turbo Pro (sin audio) o
    Kling 2.6 Pro (con audio nativo). Descuenta saldo del usuario.
    """
    FAL_KEY = os.getenv("FAL_API_KEY", "")

    # Endpoints FAL.ai por modelo
    FAL_ENDPOINTS = {
        "kling25pro": "fal-ai/kling-video/v2.5-turbo/pro/text-to-video",
        "kling26pro": "fal-ai/kling-video/v2.6/pro/text-to-video",
    }
    endpoint = FAL_ENDPOINTS.get(model_key, FAL_ENDPOINTS["kling25pro"])
    price_info = VIDEO_PRICES.get((model_key, duration_s), VIDEO_PRICES[("kling25pro", 5)])

    # ── Mejorar el prompt ──────────────────────────────────────────────────────
    async def enhance(p: str) -> str:
        try:
            msgs = [
                {"role": "system", "content": """You are an elite AI video director and prompt engineer.
Transform the user's request into a hyper-detailed, cinematic English prompt for Kling AI video models.

STRICT RULES:
1. PHOTOREALISTIC only — NEVER cartoon, CGI, animated, illustrated (unless user explicitly asks)
2. Kling excels at: fluid motion, realistic physics, cinematic camera work, sports, action, nature
3. Include ALL: camera movement (dolly, crane, tracking, handheld), lens (35mm anamorphic),
   lighting (natural, volumetric, cinematic), textures (skin, fabric, grass), subject behavior
4. Sports scenes: packed stadium crowd, worn jersey textures, ball spin physics, turf detail,
   broadcast-style camera with slow-motion replays
5. Fantasy/creative: keep lighting and physics photorealistic, only subject is fantastical
6. End with: "Photorealistic, 4K HDR, ultra-sharp, cinematic color grade, 24fps film grain"
7. Kling responds BEST to 150-200 word English prompts — be extremely specific
8. Output ONLY the enhanced prompt. No explanations, no quotes, no preamble."""},
                {"role": "user", "content": f"Video request: {p}"},
            ]
            enhanced, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")
            return enhanced.strip()[:600]
        except Exception:
            return p

    enhanced = await enhance(prompt)
    print(f"🎬 FAL.ai video: model={model_key} dur={duration_s}s endpoint={endpoint}")
    print(f"🎬 Prompt mejorado: {enhanced[:120]}...")

    if not FAL_KEY:
        return "", "❌ FAL_API_KEY no configurada en Railway. Contactá al administrador.", "error"

    # ── Verificar saldo ────────────────────────────────────────────────────────
    if user:
        credits = await get_video_credits(user)
        if round(float(credits["balance_usd"]), 4) < price_info["user_price"]:
            return (
                "",
                f"❌ Saldo insuficiente (${credits['balance_usd']:.2f} USD). "
                f"Este video cuesta ${price_info['user_price']:.2f} USD. "
                f"Cargá créditos desde el botón 🎬 en el sidebar.",
                "error · insufficient_balance"
            )

    # ── Enviar request a FAL.ai ────────────────────────────────────────────────
    fal_input = {
        "prompt": enhanced,
        "negative_prompt": (
            "cartoon, animated, CGI, 3D render, illustration, watermark, text overlay, "
            "low quality, blurry, distorted, deformed, artifacts, noise, overexposed, "
            "underexposed, bad anatomy, ugly, worst quality, jpeg artifacts"
        ),
        "duration": str(duration_s),
        "aspect_ratio": "16:9",
        "cfg_scale": 0.5,
    }
    if model_key == "kling26pro":
        fal_input["enable_audio"] = True

    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.post(
                f"https://queue.fal.run/{endpoint}",
                headers={
                    "Authorization": f"Key {FAL_KEY}",
                    "Content-Type": "application/json",
                },
                json={"input": fal_input},
            )
            d = r.json()
            print(f"🎬 FAL submit: status={r.status_code} resp={str(d)[:200]}")

            if not r.is_success:
                err = d.get("detail") or d.get("error") or f"HTTP {r.status_code}"
                return "", f"❌ Error al iniciar la generación en FAL.ai: {err}", "error"

            request_id = d.get("request_id")
            if not request_id:
                return "", "❌ FAL.ai no devolvió request_id. Intentá de nuevo.", "error"

    except Exception as e:
        print(f"🎬 FAL submit exception: {e}")
        return "", f"❌ Error de conexión con FAL.ai: {str(e)[:100]}", "error"

    # ── Polling hasta completar ────────────────────────────────────────────────
    status_url = f"https://queue.fal.run/{endpoint}/requests/{request_id}"
    video_url = None
    max_polls = 36  # 6 minutos (36 × 10s)

    async with httpx.AsyncClient(timeout=400) as c:
        for attempt in range(max_polls):
            await asyncio.sleep(10)
            try:
                r2 = await c.get(status_url, headers={"Authorization": f"Key {FAL_KEY}"})
                d2 = r2.json()
                fal_status = d2.get("status", "unknown")
                print(f"🎬 FAL polling {attempt+1}/{max_polls}: {fal_status}")

                if fal_status == "COMPLETED":
                    output = d2.get("output") or {}
                    vid = output.get("video") or {}
                    video_url = vid.get("url") if isinstance(vid, dict) else None
                    if not video_url:
                        video_url = output.get("video_url") or output.get("url")
                    print(f"✅ FAL completed: {video_url}")
                    break

                elif fal_status in ("FAILED", "ERROR", "CANCELLED"):
                    err_detail = d2.get("error") or d2.get("detail") or fal_status
                    print(f"❌ FAL failed: {err_detail}")
                    return "", f"❌ El video falló durante la generación: {str(err_detail)[:100]}", "error"

            except Exception as e:
                print(f"🎬 FAL polling exception at {attempt}: {e}")
                continue

    if not video_url:
        return (
            "",
            "❌ Timeout: el video tardó más de 6 minutos. FAL.ai puede estar con alta demanda — intentá de nuevo en unos minutos.",
            "error · timeout"
        )

    # ── Descontar saldo ────────────────────────────────────────────────────────
    if user:
        await consume_video_credit(user, model_key, duration_s, prompt[:80])

    label_model = "Kling 2.5 Pro" if model_key == "kling25pro" else "Kling 2.6 Pro (audio)"
    audio_note = " · Audio nativo incluido 🔊" if model_key == "kling26pro" else ""
    cost_note = f"\n\n> 💰 *${price_info['user_price']:.2f} USD · {label_model}{audio_note}*"

    return (
        video_url,
        f"🎬 Video generado con **FAL.ai** ({label_model} · {duration_s}s).{cost_note}",
        f"fal · {model_key} · {duration_s}s",
    )


def _friendly_video_error(raw_error: str) -> str:
    """Convierte errores técnicos de FAL en mensajes amigables para el usuario."""
    e = raw_error.lower()
    if "exhausted" in e or "locked" in e or "balance" in e or "billing" in e:
        return (
            "No puedo generar el video ahora mismo porque los créditos de video están agotados. "
            "Podés recargar desde [fal.ai/dashboard/billing](https://fal.ai/dashboard/billing) "
            "y el video quedará disponible enseguida."
        )
    if "timeout" in e:
        return "La generación del video tardó demasiado. Intentá de nuevo en unos minutos."
    if "not_pro" in e or "plan" in e:
        return "La generación de videos es exclusiva del Plan Pro."
    return f"No se pudo generar el video: {raw_error[:120]}"



# ─────────────────────────────────────────────────────────────────────────────
#  GENERACIÓN DE MÚSICA — Cascada: udioapi.pro → FAL.ai Stable Audio
# ─────────────────────────────────────────────────────────────────────────────

UDIO_API_KEY = os.getenv("UDIO_API_KEY", "")

async def generate_music_smart(prompt: str, username: str = "") -> tuple[str, str, str]:
    """
    Genera una canción completa usando cascada:
    1° udioapi.pro  (si UDIO_API_KEY está configurada y tiene créditos)
    2° FAL.ai Stable Audio (fallback automático, pay-per-use)
    Retorna: (audio_url, result_text, label)
    """

    # ── Enriquecimiento del prompt con Groq ──────────────────────────────────
    async def enhance_music_prompt(p: str) -> dict:
        try:
            msgs = [
                {"role": "system", "content": """Sos un productor musical experto.
Dado el pedido del usuario, generá:
1. Un prompt en inglés para generar la canción (descriptivo, con género, mood, instrumentos, estilo)
2. Una letra en español coherente con el pedido (2 estrofas + coro, máximo 150 palabras)
3. Tags de género (ej: "cumbia, tropical, festivo, guitarra, acordeón")
4. Un título creativo para la canción

Respondé SOLO con JSON:
{"prompt_en": "...", "lyrics": "...", "tags": "...", "title": "..."}"""},
                {"role": "user", "content": f"Pedido: {p}"}
            ]
            result, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")
            clean = result.strip().replace("```json","").replace("```","").strip()
            s, e = clean.find("{"), clean.rfind("}") + 1
            if s >= 0 and e > s:
                return json.loads(clean[s:e])
        except Exception as ex:
            print(f"music enhance error: {ex}")
        return {"prompt_en": p, "lyrics": "", "tags": "music", "title": "Mi Canción"}

    enhanced = await enhance_music_prompt(prompt)
    title  = enhanced.get("title", "Canción")
    tags   = enhanced.get("tags", "music")
    lyrics = enhanced.get("lyrics", "")
    music_prompt_en = enhanced.get("prompt_en", prompt)
    print(f"🎵 Generando música: {title} | tags: {tags}")

    # ════════════════════════════════════════════════════════════════════════
    #  OPCIÓN 1: udioapi.pro
    # ════════════════════════════════════════════════════════════════════════
    udio_skip_reason = None

    if not UDIO_API_KEY:
        udio_skip_reason = "UDIO_API_KEY no configurada"
    else:
        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.post(
                    "https://udioapi.pro/api/generate",
                    headers={
                        "Authorization": f"Bearer {UDIO_API_KEY}",
                        "Content-Type": "application/json"
                    },
                    json={
                        "prompt": music_prompt_en[:300],
                        "lyrics": lyrics[:500],
                        "tags": tags,
                        "title": title,
                        "make_instrumental": not bool(lyrics),
                    }
                )
                d = r.json()
                print(f"🎵 Udio submit: {r.status_code} | {str(d)[:200]}")

                # Detectar errores de crédito o plan → pasar al fallback
                err_msg = (d.get("message") or d.get("error") or "").lower()
                if not r.is_success or any(k in err_msg for k in ["no credit", "credit", "quota", "limit", "plan", "insufficient"]):
                    udio_skip_reason = f"Udio sin créditos ({err_msg[:60]})"
                else:
                    # Obtener task_id para polling
                    task_id = d.get("taskId") or d.get("task_id") or d.get("id")
                    if not task_id:
                        audio_url = d.get("audioUrl") or d.get("audio_url") or d.get("url")
                        if audio_url:
                            lyrics_preview = lyrics[:200]
                            lyrics_note = f"\n\n**Letra:**\n_{lyrics_preview}..._" if lyrics_preview else ""
                            result_text = f"🎵 **{title}** · _{tags}_{lyrics_note}\n\n> 🎧 *Generada con Udio AI · Hacé clic en ▶️ para escuchar*"
                            return audio_url, result_text, "udio · música"
                        udio_skip_reason = "Udio no devolvió task_id ni URL"
                    else:
                        # Polling Udio (máx 4 min)
                        audio_url = None
                        async with httpx.AsyncClient(timeout=300) as c2:
                            for attempt in range(24):
                                await asyncio.sleep(10)
                                try:
                                    r2 = await c2.get(
                                        f"https://udioapi.pro/api/task/{task_id}",
                                        headers={"Authorization": f"Bearer {UDIO_API_KEY}"}
                                    )
                                    d2 = r2.json()
                                    status = d2.get("status", "").lower()
                                    print(f"🎵 Udio polling {attempt+1}/24: {status}")

                                    if status in ("completed", "success", "done"):
                                        songs = d2.get("songs") or d2.get("tracks") or d2.get("results") or []
                                        if songs and isinstance(songs, list):
                                            audio_url = songs[0].get("audioUrl") or songs[0].get("audio_url") or songs[0].get("url")
                                        if not audio_url:
                                            audio_url = (d2.get("audioUrl") or d2.get("audio_url") or
                                                         d2.get("url") or d2.get("output"))
                                        print(f"✅ Udio completado: {audio_url}")
                                        break
                                    elif status in ("failed", "error", "cancelled"):
                                        udio_skip_reason = d2.get("error") or d2.get("message") or status
                                        break
                                except Exception as pe:
                                    print(f"🎵 Udio polling error {attempt}: {pe}")
                                    continue

                        if audio_url:
                            lyrics_preview = lyrics[:200]
                            lyrics_note = f"\n\n**Letra:**\n_{lyrics_preview}..._" if lyrics_preview else ""
                            result_text = f"🎵 **{title}** · _{tags}_{lyrics_note}\n\n> 🎧 *Generada con Udio AI · Hacé clic en ▶️ para escuchar*"
                            return audio_url, result_text, "udio · música"
                        elif not udio_skip_reason:
                            udio_skip_reason = "Timeout o sin URL en respuesta"

        except Exception as e:
            udio_skip_reason = f"Error de conexión: {str(e)[:80]}"
            print(f"🎵 Udio error: {e}")

    print(f"🎵 Udio no disponible ({udio_skip_reason}), intentando FAL.ai Stable Audio...")

    # ════════════════════════════════════════════════════════════════════════
    #  OPCIÓN 2: FAL.ai Stable Audio (fallback)
    # ════════════════════════════════════════════════════════════════════════
    FAL_KEY = os.getenv("FAL_API_KEY", "")
    fal_skip_reason = None

    if not FAL_KEY:
        fal_skip_reason = "FAL_API_KEY no configurada"
    else:
        fal_audio_url = None
        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.post(
                    "https://queue.fal.run/fal-ai/stable-audio",
                    headers={
                        "Authorization": f"Key {FAL_KEY}",
                        "Content-Type": "application/json"
                    },
                    json={
                        "prompt": music_prompt_en[:500],
                        "seconds_total": 30,
                        "steps": 100,
                    }
                )
                d = r.json()
                print(f"🎵 FAL music submit: {r.status_code} | {str(d)[:200]}")

                # Detectar saldo agotado u otros errores bloqueantes
                err_msg = (d.get("message") or d.get("detail") or "").lower()
                if not r.is_success or any(k in err_msg for k in ["locked", "exhausted", "balance", "quota", "limit", "billing"]):
                    fal_skip_reason = f"FAL sin saldo ({err_msg[:80]})"
                else:
                    request_id = d.get("request_id")
                    if not request_id:
                        fal_skip_reason = "FAL no devolvió request_id"
                    else:
                        # Polling FAL (máx 3 min)
                        async with httpx.AsyncClient(timeout=300) as c2:
                            for attempt in range(18):
                                await asyncio.sleep(10)
                                try:
                                    r2 = await c2.get(
                                        f"https://queue.fal.run/fal-ai/stable-audio/requests/{request_id}",
                                        headers={"Authorization": f"Key {FAL_KEY}"}
                                    )
                                    d2 = r2.json()
                                    status = d2.get("status", "").lower()
                                    print(f"🎵 FAL music polling {attempt+1}/18: {status}")

                                    if status in ("completed", "ok"):
                                        output = d2.get("output") or {}
                                        fal_audio_url = (
                                            output.get("audio_file", {}).get("url") or
                                            output.get("audio", {}).get("url") or
                                            output.get("url") or
                                            d2.get("audio_url")
                                        )
                                        print(f"✅ FAL music completado: {fal_audio_url}")
                                        break
                                    elif status in ("failed", "error", "cancelled"):
                                        fal_skip_reason = d2.get("error") or d2.get("detail") or status
                                        break
                                except Exception as pe:
                                    print(f"🎵 FAL polling error {attempt}: {pe}")
                                    continue

                        if fal_audio_url:
                            result_text = f"🎵 **{title}** · _{tags}_\n\n> 🎧 *Generada con FAL Stable Audio · Hacé clic en ▶️ para escuchar*"
                            return fal_audio_url, result_text, "fal · stable-audio"
                        elif not fal_skip_reason:
                            fal_skip_reason = "Timeout o sin URL en respuesta"

        except Exception as e:
            fal_skip_reason = f"Error de conexión: {str(e)[:80]}"
            print(f"🎵 FAL music error: {e}")

    print(f"🎵 FAL no disponible ({fal_skip_reason}), usando Pollinations Music (gratuito)...")

    # ════════════════════════════════════════════════════════════════════════
    #  OPCIÓN 3: Pollinations Music (siempre disponible, 100% gratuito)
    # ════════════════════════════════════════════════════════════════════════
    try:
        import urllib.parse as _urlparse
        # Pollinations Music genera audio a partir de un prompt de texto
        # Endpoint: https://text.pollinations.ai/  con model=musicgen
        poll_prompt = f"{music_prompt_en[:200]}, {tags}"
        encoded = _urlparse.quote(poll_prompt)
        # URL directa — Pollinations devuelve el audio directamente (no requiere polling)
        audio_url = f"https://audio.pollinations.ai/{encoded}"

        # Descargar el audio directamente (no confiar en HEAD — Pollinations puede responder
        # text/html en HEAD aunque el audio exista; hacemos GET y verificamos por tamaño)
        async with httpx.AsyncClient(timeout=45, follow_redirects=True) as c:
            try:
                rg = await c.get(audio_url)
                content_type = rg.headers.get("content-type", "")
                print(f"🎵 Pollinations music GET: {rg.status_code} | {content_type} | {len(rg.content)} bytes")
                if rg.is_success and len(rg.content) > 3000:
                    import uuid as _uuid
                    tok = str(_uuid.uuid4())
                    _file_cache[tok] = (rg.content, "orquesta_music.mp3", "audio/mpeg")
                    cached_url = f"/api/download/{tok}"
                    result_text = f"🎵 **{title}** · _{tags}_\n\n> 🎧 *Generada con Pollinations Music · Hacé clic en ▶️ para escuchar*"
                    return cached_url, result_text, "pollinations · music"
                elif rg.is_success:
                    result_text = f"🎵 **{title}** · _{tags}_\n\n> 🎧 *Generada con Pollinations Music · Hacé clic en ▶️ para escuchar*"
                    return audio_url, result_text, "pollinations · music"
            except Exception as ge:
                print(f"🎵 Pollinations GET error: {ge}")

    except Exception as e:
        print(f"🎵 Pollinations music error: {e}")

    # Si los 3 fallaron, mensaje claro al usuario
    return (
        "",
        f"❌ No pude generar la música ahora mismo:\n"
        f"- Udio: {udio_skip_reason}\n"
        f"- FAL: {fal_skip_reason}\n"
        f"- Pollinations: no disponible\n\n"
        f"Recargá saldo en [fal.ai/dashboard/billing](https://fal.ai/dashboard/billing) o intentá de nuevo en unos minutos.",
        "error"
    )

# ─────────────────────────────────────────────────────────────────────────────
#  DEEP SEARCH
# ─────────────────────────────────────────────────────────────────────────────

async def deep_search(query: str) -> str:
    results = []

    if TAVILY_KEY:
        try:
            ctx = await call_tavily(query, depth="advanced", max_results=8)
            if ctx: results.append(f"=== WEB (TIEMPO REAL) ===\n{ctx}")
        except: pass

    try:
        import urllib.parse
        term = urllib.parse.quote(query[:60])
        async with httpx.AsyncClient(timeout=10) as c:
            r = await c.get(f"https://es.wikipedia.org/api/rest_v1/page/summary/{term}",
                headers={"User-Agent":"OrquestaAI/1.0"})
            if r.is_success:
                d = r.json()
                extract = d.get("extract","")
                if extract and len(extract) > 100:
                    results.append(f"=== WIKIPEDIA: {d.get('title','')} ===\n{extract[:800]}")
    except: pass

    return "\n\n".join(results) if results else ""

# ─────────────────────────────────────────────────────────────────────────────
#  FILE GENERATION
# ─────────────────────────────────────────────────────────────────────────────

from app.file_generator import generate_excel, generate_docx, generate_pdf, FILE_SYSTEM_PROMPTS

_file_cache: dict = {}  # token -> (bytes, filename, mime, timestamp)

def cache_file(file_bytes, filename, mime):
    token = str(uuid.uuid4())
    _file_cache[token] = (file_bytes, filename, mime, time.time())
    # Limpiar archivos de más de 2 horas o si hay más de 50
    now = time.time()
    expired = [k for k, v in _file_cache.items() if now - v[3] > 7200]
    for k in expired:
        del _file_cache[k]
    if len(_file_cache) > 50:
        oldest = sorted(_file_cache.keys(), key=lambda k: _file_cache[k][3])[:10]
        for k in oldest:
            del _file_cache[k]
    return token

def extract_title(prompt):
    stop = r'\b(excel|word|pdf|planilla|documento|genera|crea|haceme|hacé|un|una|el|la|en|como|formato|me|por|favor|quiero|necesito|dame|reporte|informe)\b'
    clean = re.sub(stop, '', prompt.lower(), flags=re.IGNORECASE)
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean.title()[:60] or "Documento Orquesta"

async def generate_file_from_prompt(prompt, file_type, username=""):
    file_system = FILE_SYSTEM_PROMPTS.get(file_type, FILE_SYSTEM_PROMPTS["pdf"])
    if username: file_system += f"\n\nEl usuario se llama {username}."
    
    # Enriquecer el prompt para mayor calidad
    enhanced_prompt = f"""Generá contenido de MÁXIMA CALIDAD PROFESIONAL para el siguiente pedido.
El contenido debe ser:
- Completo y detallado (no básico)
- Con datos reales, ejemplos concretos y estructura profesional
- Formateado correctamente para {file_type.upper()}
- Listo para usar en un contexto real de trabajo o negocio

Pedido: {prompt}

Incluí tantos datos, filas, columnas, secciones o contenido como sea necesario para que el archivo sea verdaderamente útil y completo."""
    
    msgs = [{"role":"system","content":file_system},{"role":"user","content":enhanced_prompt}]
    # Usar solo Groq para archivos — Gemini no es necesario y puede estar agotado
    ai_content, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile", use_gemini_fallback=False)
    title = extract_title(prompt)
    safe = title[:30].replace(' ','_').replace('/','').replace('\\','')
    if file_type == "xlsx":
        fb = generate_excel(ai_content, title); fn = f"orquesta_{safe}.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif file_type == "docx":
        fb = generate_docx(ai_content, title); fn = f"orquesta_{safe}.docx"
        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    else:
        fb = generate_pdf(ai_content, title); fn = f"orquesta_{safe}.pdf"
        mime = "application/pdf"
    return fb, fn, mime


# ─────────────────────────────────────────────────────────────────────────────
#  MODELOS
# ─────────────────────────────────────────────────────────────────────────────

class OrchestrateReq(BaseModel):
    prompt: str
    history: list = []
    mode: str = "general"
    username: str = ""
    language: str = ""
    tts_enabled: bool = False
    conversation_id: str = ""
    user_id: str = ""           # id de tabla users
    auth_id: str = ""           # auth_id de Supabase Auth
    user_plan: str = ""         # ignorado por seguridad — el plan siempre viene de DB
    # Opciones de video
    video_model: str = "kling25pro"   # kling25pro | kling26pro
    video_duration: int = 5           # 5 | 10

class OrchestrateResp(BaseModel):
    result: str
    task_type: str
    model_label: str
    latency_ms: int
    image_url: str = ""
    file_url: str = ""
    file_type: str = ""
    file_name: str = ""
    tts_url: str = ""
    video_url: str = ""
    music_url: str = ""         # URL directa de audio generado
    is_pro: bool = False
    daily_remaining: int = -1
    video_credits_remaining: int = -1
    upgrade_banner: str = ""
    upgrade_cta: str = ""
    upgrade_cta_url: str = ""


# ─────────────────────────────────────────────────────────────────────────────
#  ENDPOINT PRINCIPAL: /api/orchestrate
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/orchestrate", response_model=OrchestrateResp)
async def orchestrate(req: OrchestrateReq, request: Request, background_tasks: BackgroundTasks, authorization: str = Header(None)):
    if not req.prompt.strip(): raise HTTPException(400, "Prompt vacío")
    # Rate limiting por IP
    client_ip = request.client.host if request.client else "unknown"
    if not check_rate_limit(client_ip, max_per_minute=30):
        raise HTTPException(429, "Demasiadas peticiones. Esperá un momento.")

    user = await get_optional_user(authorization)
    print(f"orchestrate: auth_id='{req.auth_id[:8] if req.auth_id else ''}' user_id='{req.user_id[:8] if req.user_id else ''}' user_plan='{req.user_plan}' user={'found' if user else 'None'}")

    # Fallback: buscar usuario via REST API de Supabase con httpx async
    if not user and SUPABASE_URL and SUPABASE_SERVICE_KEY:
        # Extraer auth_id del JWT sin verificar firma
        jwt_auth_id = ""
        if authorization and authorization.startswith("Bearer "):
            try:
                token = authorization.replace("Bearer ", "").strip()
                parts = token.split(".")
                if len(parts) == 3:
                    pad = parts[1] + "=" * (4 - len(parts[1]) % 4)
                    jwt_payload = json.loads(base64.urlsafe_b64decode(pad))
                    jwt_auth_id = jwt_payload.get("sub", "")
            except Exception:
                pass

        # Intentar con todos los IDs disponibles
        for field, value in [
            ("auth_id", jwt_auth_id),
            ("auth_id", req.auth_id),
            ("auth_id", req.user_id),
            ("id", req.user_id),
        ]:
            if not value:
                continue
            try:
                async with httpx.AsyncClient(timeout=10) as hx:
                    url = f"{SUPABASE_URL}/rest/v1/users?{field}=eq.{value}&select=*&limit=1"
                    resp = await hx.get(url, headers={
                        "apikey": SUPABASE_SERVICE_KEY,
                        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                    })
                    if resp.is_success:
                        data_list = resp.json()
                        if data_list:
                            user = data_list[0]
                            print(f"✅ Usuario via REST: {field}={value[:8]} plan={user.get('plan')}")
                            break
            except Exception as e:
                print(f"REST fallback {field}: {e}")
                continue

    username = user["name"] if user else req.username

    t0 = time.time()
    task = classify(req.prompt, req.mode, req.history)
    label = TASK_LABELS.get(task, "groq · llama 3.3")

    # ── Verificar acceso ──────────────────────────────────────────────────────
    block = check_pro_access(user, task)
    if block and block.get("blocked"):
        return OrchestrateResp(
            result=block["message"],
            task_type=task,
            model_label="orquesta · plan",
            latency_ms=int((time.time()-t0)*1000),
            is_pro=False,
            daily_remaining=0,
            upgrade_banner=block["message"],
            upgrade_cta=block.get("cta","Activar Pro"),
            upgrade_cta_url=block.get("cta_url","/pricing"),
        )

    # ── Incrementar contador de mensajes ──────────────────────────────────────
    if user and supabase:
        try:
            supabase.rpc("increment_message_count", {"p_user_id": user["id"]}).execute()
        except: pass

    # ── Detectar feedback de errores del usuario y registrarlos ───────────────
    feedback_triggers = [
        "detecté", "detecte", "encontré", "encontre", "hay un error",
        "está fallando", "esta fallando", "no funciona", "analices estos errores",
        "propone una mejora", "no son premium", "mala calidad", "mejorar",
        "autorepar", "auto-mejor", "self-improv"
    ]
    is_feedback = False
    prompt_lower = req.prompt.lower()
    if any(k in prompt_lower for k in feedback_triggers):
        log_error("/chat/feedback", f"Usuario reportó: {req.prompt[:150]}", "user_feedback")
        is_feedback = True
        print(f"📝 Feedback registrado — se enviará análisis en background")

    img_url = file_url = file_type = file_name = tts_url = result = video_url = ""
    label = "orquesta · llama 3.3"   # default siempre definido
    system = get_system(req.mode, username, req.history)
    is_pro = user.get("plan") == "pro" if user else False
    daily_remaining = -1
    if user and not is_pro:
        daily_remaining = max(0, FREE_DAILY_LIMIT - user.get("daily_message_count", 0) - 1)

    try:
        if task.startswith("file_gen_"):
            ftype = task.replace("file_gen_","")
            try:
                fb, fn, mime = await generate_file_from_prompt(req.prompt, ftype, username)
                if not fb or len(fb) < 100:
                    raise Exception("El archivo generado está vacío")
                token = cache_file(fb, fn, mime)
                file_url = f"/api/download/{token}"
                file_type = ftype; file_name = fn
                names = {"xlsx":"Excel","docx":"Word","pdf":"PDF"}
                result = f"✅ Tu archivo **{names.get(ftype,ftype.upper())}** está listo. Hacé clic en el botón para descargarlo."
                label = f"orquesta · {names.get(ftype,ftype).lower()}"
            except Exception as e:
                err_detail = str(e) if str(e) else type(e).__name__
                print(f"File gen error ({ftype}): {err_detail}")
                # Detectar si es error de quota de Gemini
                if "429" in err_detail or "quota" in err_detail.lower() or "exhausted" in err_detail.lower():
                    result = (f"Tuve un problema técnico generando el archivo — una API de soporte está "
                             f"temporalmente saturada. Intentá de nuevo en unos minutos o reformulá el pedido.")
                else:
                    # Intentar con prompt simplificado
                    try:
                        simple_prompt = f"Creá un {ftype} básico y completo sobre: {req.prompt[:150]}"
                        fb2, fn2, mime2 = await generate_file_from_prompt(simple_prompt, ftype, username)
                        token2 = cache_file(fb2, fn2, mime2)
                        file_url = f"/api/download/{token2}"
                        file_type = ftype; file_name = fn2
                        names = {"xlsx":"Excel","docx":"Word","pdf":"PDF"}
                        result = f"✅ Tu archivo **{names.get(ftype,ftype.upper())}** está listo."
                        label = f"orquesta · {names.get(ftype,ftype).lower()}"
                    except Exception as e2:
                        result = f"No pude generar el archivo. Reformulá el pedido con más detalle."

        elif task == "video_gen":
            video_url, result, label = await generate_video_smart(
                req.prompt, req.history, req.mode, username,
                user=user,
                model_key=req.video_model,
                duration_s=req.video_duration,
            )
            # Convertir error técnico en mensaje amigable
            if label == "error" and not video_url:
                result = _friendly_video_error(result)

        elif task == "image_gen":
            img_url, result, label = await generate_image_smart(req.prompt)

        elif task == "realtime":
            try:
                ctx = await deep_search(req.prompt)
                if ctx:
                    synth = (f'El usuario pregunta: "{req.prompt}"\n\n'
                             f"=== INFORMACIÓN DE MÚLTIPLES FUENTES ===\n{ctx}\n\n"
                             f"Respondé de forma COMPLETA Y DETALLADA integrando todos los datos.")
                    msgs = build_messages(system, req.history[:-1], synth)
                    result, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")
                    label = "orquesta · web"
                else:
                    msgs = build_messages(system, req.history, req.prompt)
                    result, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")
            except Exception:
                msgs = build_messages(system, req.history, req.prompt)
                result, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")

        elif task == "sound_gen":
            music_url, result, label = await generate_music_smart(req.prompt, username)
            if music_url:
                # Guardar en caché para descarga
                import urllib.request as _ur
                try:
                    audio_resp = _ur.urlopen(music_url, timeout=30)
                    audio_bytes = audio_resp.read()
                    audio_token = cache_file(audio_bytes, "orquesta_music.mp3", "audio/mpeg")
                    tts_url = f"/api/download/{audio_token}"
                except Exception as _me:
                    print(f"Music cache error: {_me}")
                    tts_url = music_url  # usar URL directa si falla el cache

        elif task == "translate":
            ts = system + "\n\nEres un traductor experto. Traducí con precisión y naturalidad."
            msgs = build_messages(ts, req.history, req.prompt)
            result, _ = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")
            label = "groq · traductor"

        elif task == "analysis" and TAVILY_KEY:
            try:
                ctx = await deep_search(req.prompt)
                if ctx:
                    enriched = (f'"{req.prompt}"\n\nDatos reales de internet:\n{ctx[:3000]}\n\n'
                                f"Realizá un análisis PROFUNDO Y COMPARATIVO con datos reales.")
                    msgs = build_messages(system, req.history, enriched)
                    result, _ = await groq_with_fallback(msgs, "llama-3.1-8b-instant")
                    label = "orquesta · análisis+web"
                else:
                    msgs = build_messages(system, req.history, req.prompt)
                    result, _ = await groq_with_fallback(msgs, "llama-3.1-8b-instant")
            except Exception:
                msgs = build_messages(system, req.history, req.prompt)
                result, _ = await groq_with_fallback(msgs, "llama-3.1-8b-instant")

        else:
            model = TASK_MODELS.get(task, "llama-3.3-70b-versatile")
            msgs = build_messages(system, req.history, req.prompt)
            result, used = await groq_with_fallback(msgs, model)
            if used == "gemini": label = "gemini · flash"

        # ── TTS (solo Pro) ────────────────────────────────────────────────────
        # TTS solo si hay resultado real de texto (no si es descripción de video fallido)
        skip_tts = task == "video_gen" and not video_url  # No TTS cuando video falló
        if req.tts_enabled and is_pro and result and len(result) < 8000 and not file_url and not skip_tts:
            try:
                clean_text = re.sub(r'```[\s\S]*?```', '', result)
                clean_text = re.sub(r'[#*`_~>]', '', clean_text)
                clean_text = re.sub(r'\s+', ' ', clean_text).strip()[:2000]
                audio_bytes = await call_openai_tts(clean_text, voice="nova")
                audio_token = str(uuid.uuid4())
                _file_cache[audio_token] = (audio_bytes, "response.mp3", "audio/mpeg")
                tts_url = f"/api/download/{audio_token}"
            except Exception:
                pass

        # ── Guardar en Supabase ────────────────────────────────────────────────
        # Solo guardar en Supabase si el conversation_id es un UUID válido (no sesión local)
        is_valid_uuid = req.conversation_id and not req.conversation_id.startswith("session_") and len(req.conversation_id) == 36
        if user and is_valid_uuid and supabase:
            try:
                supabase.table("messages").insert({
                    "conversation_id": req.conversation_id,
                    "role": "user",
                    "content": req.prompt,
                }).execute()
                supabase.table("messages").insert({
                    "conversation_id": req.conversation_id,
                    "role": "assistant",
                    "content": result,
                    "model_label": label,
                    "task_type": task,
                    "latency_ms": int((time.time()-t0)*1000),
                    "has_image": bool(img_url),
                    "has_file": bool(file_url),
                    "has_video": bool(video_url),
                }).execute()
                supabase.table("conversations").update({
                    "title": req.prompt[:60],
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }).eq("id", req.conversation_id).execute()
            except Exception as e:
                print(f"Warning: no se pudo guardar en Supabase: {e}")

    except HTTPException: raise
    except Exception as e: raise HTTPException(502, detail=str(e))

    # Consultar créditos de video si se usó video_gen
    video_credits_remaining = -1
    if task == "video_gen" and user:
        try:
            vc = await get_video_credits(user)
            video_credits_remaining = vc.get("available", -1)
        except Exception:
            pass

    # music_url: si sound_gen tuvo éxito, está en tts_url (caché local)
    # Si falló el caché, music_url tiene la URL directa
    _music_url = music_url if task == "sound_gen" and "music_url" in dir() else ""

    # Disparar análisis de auto-mejoramiento DESPUÉS de responder al usuario
    if is_feedback:
        background_tasks.add_task(analyze_and_propose_improvements)

    return OrchestrateResp(
        result=result, task_type=task, model_label=label,
        latency_ms=int((time.time()-t0)*1000),
        image_url=img_url, file_url=file_url, file_type=file_type,
        file_name=file_name, tts_url=tts_url, video_url=video_url,
        music_url=_music_url,
        is_pro=is_pro,
        daily_remaining=daily_remaining,
        video_credits_remaining=video_credits_remaining,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  AUTH ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    is_pro = (
        user.get("plan") == "pro" and (
            user.get("plan_expires_at") is None or
            parse_expiry(user["plan_expires_at"]) > datetime.now(timezone.utc)
        )
    )
    return {
        "id":             user["id"],
        "email":          user["email"],
        "name":           user["name"],
        "avatar_url":     user["avatar_url"],
        "plan":           user["plan"],
        "is_pro":         is_pro,
        "plan_expires_at": user.get("plan_expires_at"),
        "daily_message_count": user.get("daily_message_count", 0),
        "daily_remaining": max(0, FREE_DAILY_LIMIT - user.get("daily_message_count", 0)) if not is_pro else -1,
        "free_daily_limit": FREE_DAILY_LIMIT,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  CONVERSACIONES ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/conversations")
async def list_conversations(user: dict = Depends(get_current_user)):
    if not supabase: raise HTTPException(500, "Supabase no configurado")
    limit = 50 if user.get("plan") == "pro" else 3
    result = supabase.table("conversations")\
        .select("id, title, mode, created_at, updated_at")\
        .eq("user_id", user["id"])\
        .order("updated_at", desc=True)\
        .limit(limit)\
        .execute()
    return {"conversations": result.data or [], "limit": limit, "is_pro": user.get("plan") == "pro"}

@router.post("/conversations")
async def create_conversation(data: dict, user: dict = Depends(get_current_user)):
    if not supabase: raise HTTPException(500, "Supabase no configurado")
    result = supabase.table("conversations").insert({
        "user_id": user["id"],
        "title":   data.get("title", "Nueva conversación"),
        "mode":    data.get("mode", "general"),
    }).execute()
    return result.data[0] if result.data else {}

@router.get("/conversations/{conv_id}/messages")
async def get_messages(conv_id: str, user: dict = Depends(get_current_user)):
    if not supabase: raise HTTPException(500, "Supabase no configurado")
    conv = supabase.table("conversations")\
        .select("id")\
        .eq("id", conv_id)\
        .eq("user_id", user["id"])\
        .single().execute()
    if not conv.data:
        raise HTTPException(404, "Conversación no encontrada")
    messages = supabase.table("messages")\
        .select("*")\
        .eq("conversation_id", conv_id)\
        .order("created_at")\
        .execute()
    return {"messages": messages.data or []}

@router.delete("/conversations/{conv_id}")
async def delete_conversation(conv_id: str, user: dict = Depends(get_current_user)):
    if not supabase: raise HTTPException(500, "Supabase no configurado")
    supabase.table("conversations")\
        .delete()\
        .eq("id", conv_id)\
        .eq("user_id", user["id"])\
        .execute()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
#  CHECKOUT ENDPOINTS — FIX 3: sin Depends, con fallback por user_id
# ─────────────────────────────────────────────────────────────────────────────

class CheckoutReq(BaseModel):
    plan: str                  # "pro_monthly" | "pro_annual"
    provider: str = "mercadopago"
    user_id: str = ""          # enviado desde el frontend como fallback
    user_email: str = ""       # enviado desde el frontend como fallback
    access_token: str = ""

@router.post("/checkout")
async def create_checkout(req: CheckoutReq, authorization: str = Header(None)):
    """Crea una sesión de pago (MercadoPago o Stripe)."""

    # 1. Intentar obtener usuario por JWT
    user = await get_optional_user(authorization)

    # 2. Fallback: buscar por user_id en Supabase si el JWT falló
    if not user and req.user_id and supabase:
        try:
            result = supabase.table("users").select("*").eq("id", req.user_id).single().execute()
            if result.data:
                user = result.data
        except:
            pass

    # 3. Último fallback: construir user mínimo con datos del body
    if not user and req.user_email:
        user = {
            "id":       req.user_id or str(uuid.uuid4()),
            "email":    req.user_email,
            "name":     "",
            "auth_id":  "",
        }

    if not user:
        raise HTTPException(401, "No autenticado — iniciá sesión para suscribirte")

    if req.provider == "mercadopago":
        return await _mp_checkout(req.plan, user)
    elif req.provider == "stripe":
        return await _stripe_checkout(req.plan, user)
    else:
        raise HTTPException(400, "Proveedor inválido")


async def _stripe_checkout(plan: str, user: dict) -> dict:
    if not STRIPE_SECRET_KEY:
        raise HTTPException(500, "Stripe no configurado")

    price_id = STRIPE_PRICE_MONTHLY if plan == "pro_monthly" else STRIPE_PRICE_ANNUAL
    if not price_id:
        raise HTTPException(500, f"Price ID para {plan} no configurado en Railway")

    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post(
            "https://api.stripe.com/v1/checkout/sessions",
            auth=(STRIPE_SECRET_KEY, ""),
            data={
                "mode": "subscription",
                "line_items[0][price]": price_id,
                "line_items[0][quantity]": "1",
                "customer_email": user["email"],
                "client_reference_id": user["id"],
                "metadata[auth_id]": user.get("auth_id", ""),
                "metadata[plan]": plan,
                "success_url": f"{APP_URL}/?checkout=success&plan={plan}",
                "cancel_url":  f"{APP_URL}/pricing?checkout=cancelled",
                "allow_promotion_codes": "true",
            }
        )
        d = r.json()
        if not r.is_success:
            raise HTTPException(502, f"Error Stripe: {d.get('error',{}).get('message','Unknown')}")
        return {"checkout_url": d["url"], "session_id": d["id"]}


async def _mp_checkout(plan: str, user: dict) -> dict:
    if not MP_ACCESS_TOKEN:
        raise HTTPException(500, "MercadoPago no configurado — agregá MERCADOPAGO_ACCESS_TOKEN en Railway")

    prices = {
        "pro_monthly": {"title": "Orquesta Pro — Mensual", "price": 9, "currency": "USD"},
        "pro_annual":  {"title": "Orquesta Pro — Anual",   "price": 95, "currency": "USD"},
    }
    p = prices.get(plan, prices["pro_monthly"])

    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post(
            "https://api.mercadopago.com/checkout/preferences",
            headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json={
                "items": [{
                    "title":       p["title"],
                    "quantity":    1,
                    "unit_price":  p["price"],
                    "currency_id": p["currency"],
                }],
                "payer": {"email": user["email"]},
                "external_reference": user["id"],
                "metadata": {"auth_id": user.get("auth_id", ""), "plan": plan},
                "back_urls": {
                    "success": f"{APP_URL}/?checkout=success&plan={plan}",
                    "failure": f"{APP_URL}/pricing?checkout=failed",
                    "pending": f"{APP_URL}/pricing?checkout=pending",
                },
                "auto_return": "approved",
                "notification_url": f"{APP_URL}/api/webhooks/mercadopago",
            }
        )
        d = r.json()
        if not r.is_success:
            raise HTTPException(502, f"Error MercadoPago: {d.get('message','Unknown')}")

        return {
            "checkout_url": d.get("init_point") or d.get("sandbox_init_point"),
            "preference_id": d.get("id"),
        }


# ─────────────────────────────────────────────────────────────────────────────
#  WEBHOOKS
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/webhooks/stripe")
async def stripe_webhook(request: Request):
    if not STRIPE_SECRET_KEY:
        raise HTTPException(500, "Stripe no configurado")

    payload = await request.body()
    sig_header = request.headers.get("stripe-signature", "")

    try:
        import stripe as stripe_lib
        stripe_lib.api_key = STRIPE_SECRET_KEY
        event = stripe_lib.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
    except Exception as e:
        raise HTTPException(400, f"Webhook inválido: {str(e)}")

    event_type = event["type"]
    data = event["data"]["object"]

    if event_type == "checkout.session.completed":
        auth_id  = data.get("metadata", {}).get("auth_id", "")
        plan     = data.get("metadata", {}).get("plan", "pro_monthly")
        customer = data.get("customer", "")
        sub_id   = data.get("subscription", "")

        if auth_id and supabase:
            try:
                supabase.rpc("activate_pro_plan", {
                    "p_auth_id":         auth_id,
                    "p_plan_type":       plan,
                    "p_stripe_customer": customer,
                    "p_stripe_sub":      sub_id,
                }).execute()
                supabase.table("payment_events").insert({
                    "provider":   "stripe",
                    "event_type": event_type,
                    "event_id":   event.get("id"),
                    "amount":     data.get("amount_total"),
                    "currency":   data.get("currency", "usd").upper(),
                    "plan":       plan,
                    "status":     "success",
                    "raw_payload": dict(data),
                }).execute()
            except Exception as e:
                print(f"Error activando Pro en Supabase: {e}")

    elif event_type == "customer.subscription.deleted":
        customer = data.get("customer", "")
        if customer and supabase:
            try:
                user_result = supabase.table("users")\
                    .select("auth_id")\
                    .eq("stripe_customer_id", customer)\
                    .single().execute()
                if user_result.data:
                    supabase.table("users").update({
                        "plan": "free",
                        "plan_expires_at": datetime.now(timezone.utc).isoformat(),
                    }).eq("stripe_customer_id", customer).execute()
            except Exception as e:
                print(f"Error degradando a Free: {e}")

    return {"received": True}


@router.post("/webhooks/mercadopago")
async def mercadopago_webhook(request: Request):
    if not MP_ACCESS_TOKEN:
        return {"received": True}

    try:
        body = await request.json()
    except:
        return {"received": True}

    topic = body.get("type") or request.query_params.get("topic", "")
    resource_id = body.get("data", {}).get("id") or request.query_params.get("id", "")

    if topic == "payment" and resource_id:
        try:
            async with httpx.AsyncClient(timeout=15) as c:
                r = await c.get(
                    f"https://api.mercadopago.com/v1/payments/{resource_id}",
                    headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
                )
                payment = r.json()

            if payment.get("status") == "approved":
                metadata = payment.get("metadata", {})
                auth_id  = metadata.get("auth_id", "")
                plan     = metadata.get("plan", "pro_monthly")
                ext_ref  = payment.get("external_reference", "")

                if (auth_id or ext_ref) and supabase:
                    if not auth_id:
                        user_r = supabase.table("users").select("auth_id").eq("id", ext_ref).single().execute()
                        if user_r.data: auth_id = user_r.data["auth_id"]

                    if auth_id:
                        supabase.rpc("activate_pro_plan", {
                            "p_auth_id":   auth_id,
                            "p_plan_type": plan,
                            "p_mp_sub":    str(resource_id),
                        }).execute()

                        supabase.table("payment_events").insert({
                            "provider":   "mercadopago",
                            "event_type": "payment.approved",
                            "event_id":   str(resource_id),
                            "amount":     int(payment.get("transaction_amount", 0) * 100),
                            "currency":   payment.get("currency_id", "ARS"),
                            "plan":       plan,
                            "status":     "success",
                            "raw_payload": payment,
                        }).execute()
        except Exception as e:
            print(f"Error webhook MP: {e}")

    return {"received": True}


# ─────────────────────────────────────────────────────────────────────────────
#  ENDPOINTS EXISTENTES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/debug-config")
async def debug_config():
    return {
        "groq":          bool(GROQ_KEY),
        "openai":        bool(OPENAI_KEY),
        "gemini":        bool(GEMINI_KEY),
        "tavily":        bool(TAVILY_KEY),
        "supabase":      bool(SUPABASE_URL),
        "stripe":        bool(STRIPE_SECRET_KEY),
        "mercadopago":   bool(MP_ACCESS_TOKEN),
        "videogen_url":  bool(os.getenv("VIDEOGEN_URL","")),
        "status":        "ok"
    }

@router.get("/download/{token}")
async def download_file(token: str):
    if token not in _file_cache: raise HTTPException(404, "Archivo no encontrado o expirado (máx 2 horas)")
    cached = _file_cache[token]
    file_bytes, filename, mime = cached[0], cached[1], cached[2]
    return Response(content=file_bytes, media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.post("/upload")
async def upload_file(
    prompt: str = Form(default="Analizá este archivo en detalle."),
    file: UploadFile = File(...),
    username: str = Form(default=""),
    mode: str = Form(default="general"),
    authorization: str = Header(None),
):
    user = await get_optional_user(authorization)

    t0 = time.time(); raw = await file.read()
    fname = (file.filename or "").lower(); mime_type = file.content_type or ""
    uname = user["name"] if user else username
    system = get_system(mode, uname)

    async def groq_analyze(text):
        fp = f"Archivo: {file.filename}\n\nContenido:\n{text[:14000]}\n\nConsulta: {prompt}"
        msgs = build_messages(system, [], fp)
        result, used = await groq_with_fallback(msgs, "llama-3.3-70b-versatile")
        return result, ("gemini · flash" if used == "gemini" else "groq · llama 3.3")

    result = ""; label = "orquesta"; img_url = ""
    try:
        is_image = mime_type.startswith("image/") or any(fname.endswith(x) for x in [".jpg",".jpeg",".png",".gif",".webp",".bmp"])
        is_edit = is_image and any(k in prompt.lower() for k in IMAGE_EDIT_KW)

        if is_image and is_edit:
            # Flujo 1: OpenAI (GPT-4o + DALL-E 3)
            if OPENAI_KEY:
                try:
                    img_url = await call_openai_image_edit(raw, prompt)
                    result = "✅ Imagen editada con GPT-4o + DALL-E 3."
                    label = "gpt-4o + dall-e-3"
                except Exception as e:
                    result = f"Error al editar: {str(e)[:200]}"
                    label = "error"
            # Flujo 2: Gemini Vision analiza la imagen → Pollinations genera la edición
            elif GEMINI_KEY:
                try:
                    import urllib.parse as _urlparse
                    b64 = base64.b64encode(raw).decode()
                    analysis_prompt = (
                        f"Analyze this image in detail and write a single English prompt (max 200 words) "
                        f"to recreate it WITH this modification applied: {prompt}. "
                        f"Reply ONLY with the prompt text, no explanations."
                    )
                    new_prompt = await call_gemini_vision(analysis_prompt, b64, mime_type or "image/jpeg")
                    new_prompt = new_prompt.strip()[:400]
                    seed = int(time.time()) % 99999
                    encoded = _urlparse.quote(new_prompt)
                    img_url = (
                        f"https://image.pollinations.ai/prompt/{encoded}"
                        f"?width=1024&height=1024&nologo=true&enhance=true&seed={seed}&model=flux&nofeed=true"
                    )
                    result = "✅ Imagen editada con Gemini Vision + Pollinations Flux."
                    label = "gemini · vision + pollinations"
                except Exception as e:
                    result = f"Error al editar la imagen: {str(e)[:200]}"
                    label = "error"
            else:
                result = "Para editar imágenes configurá OPENAI_API_KEY o GEMINI_API_KEY."
        elif is_image:
            if GEMINI_KEY:
                b64 = base64.b64encode(raw).decode()
                result = await call_gemini_vision(f"{system}\n\nAnalizá: {prompt}", b64, mime_type or "image/jpeg")
                label = "gemini · visión"
            else: result = "Para analizar imágenes configurá GEMINI_API_KEY."
        elif fname.endswith(".pdf") or mime_type == "application/pdf":
            if GEMINI_KEY:
                b64 = base64.b64encode(raw).decode()
                result = await call_gemini_vision(f"{system}\n\nAnalizá este PDF: {prompt}", b64, "application/pdf")
                label = "gemini · pdf"
            else: result = "Para leer PDFs configurá GEMINI_API_KEY."
        elif fname.endswith(".docx"):
            try:
                from docx import Document
                doc = Document(io.BytesIO(raw))
                paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
                text = "\n".join(paras)
                result, label = await groq_analyze(text)
            except Exception as e: result = f"No se pudo leer el Word: {e}"; label = "error"
        elif any(fname.endswith(x) for x in [".xlsx",".xls"]):
            try:
                import openpyxl; wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                sheets = []
                for sname in wb.sheetnames[:6]:
                    ws = wb[sname]; rows = []
                    for row in ws.iter_rows(max_row=300, values_only=True):
                        cells = [str(c) for c in row if c is not None]
                        if cells: rows.append(" | ".join(cells))
                    if rows: sheets.append(f"=== {sname} ===\n" + "\n".join(rows[:150]))
                result, label = await groq_analyze("\n\n".join(sheets))
            except Exception as e: result = f"No se pudo leer el Excel: {e}"; label = "error"
        elif any(fname.endswith(x) for x in [".txt",".md",".csv",".json",".xml",".py",".js",".ts",".html",".css",".sql",".yaml",".yml",".log",".sh"]):
            result, label = await groq_analyze(raw.decode("utf-8", errors="ignore"))
        else: result = f"Formato no soportado: {fname}"
    except Exception as e: raise HTTPException(502, detail=f"Error: {e}")

    return {"result":result,"task_type":"file","model_label":label,
            "latency_ms":int((time.time()-t0)*1000),"image_url":img_url,"filename":file.filename}

@router.post("/tts")
async def text_to_speech(data: dict, authorization: str = Header(None)):
    if not OPENAI_KEY: raise HTTPException(400, "OPENAI_API_KEY no configurada")
    user = await get_optional_user(authorization)
    if user and user.get("plan") != "pro":
        raise HTTPException(403, "TTS disponible solo en plan Pro")
    text = data.get("text",""); voice = data.get("voice","nova")
    if not text: raise HTTPException(400, "Texto vacío")
    try:
        clean = re.sub(r'```[\s\S]*?```', '', text)
        clean = re.sub(r'[#*`_~>]', '', clean)
        clean = re.sub(r'\s+', ' ', clean).strip()[:5000]
        audio = await call_openai_tts(clean, voice)
        return StreamingResponse(io.BytesIO(audio), media_type="audio/mpeg")
    except Exception as e: raise HTTPException(502, str(e))

@router.post("/stt")
async def speech_to_text(file: UploadFile = File(...)):
    try:
        audio_bytes = await file.read()
        fname = file.filename or "audio.webm"
        content_type = file.content_type or "audio/webm"

        # Normalizar extensión
        if "ogg" in content_type:
            fname, content_type = "audio.ogg", "audio/ogg"
        elif "mp4" in content_type or "m4a" in content_type:
            fname, content_type = "audio.mp4", "audio/mp4"
        else:
            fname, content_type = "audio.webm", "audio/webm"

        print(f"STT: {fname} {len(audio_bytes)} bytes")

        # 1. Intentar con Groq Whisper (gratis)
        if GROQ_KEY:
            try:
                async with httpx.AsyncClient(timeout=60) as c:
                    r = await c.post(
                        "https://api.groq.com/openai/v1/audio/transcriptions",
                        headers={"Authorization": f"Bearer {GROQ_KEY}"},
                        files={"file": (fname, audio_bytes, content_type)},
                        data={"model": "whisper-large-v3-turbo", "language": "es", "response_format": "json"}
                    )
                    d = r.json()
                    print(f"Groq STT: {r.status_code} | {str(d)[:100]}")
                    if r.is_success and d.get("text"):
                        return {"text": d["text"].strip()}
            except Exception as e:
                print(f"Groq STT error: {e}")

        # 2. Fallback: OpenAI Whisper
        if OPENAI_KEY:
            try:
                async with httpx.AsyncClient(timeout=60) as c:
                    r = await c.post(
                        "https://api.openai.com/v1/audio/transcriptions",
                        headers={"Authorization": f"Bearer {OPENAI_KEY}"},
                        files={"file": (fname, audio_bytes, content_type)},
                        data={"model": "whisper-1", "language": "es"}
                    )
                    d = r.json()
                    if r.is_success and d.get("text"):
                        return {"text": d["text"].strip()}
                    raise Exception(d.get("error", {}).get("message", str(d)))
            except Exception as e:
                print(f"OpenAI STT error: {e}")
                raise HTTPException(502, str(e))

        raise HTTPException(400, "No hay API key de STT configurada (GROQ_API_KEY o OPENAI_API_KEY)")
    except HTTPException:
        raise
    except Exception as e:
        print(f"STT error: {e}")
        raise HTTPException(502, str(e))



# ─────────────────────────────────────────────────────────────────────────────
#  SISTEMA DE AUTO-MEJORAMIENTO
# ─────────────────────────────────────────────────────────────────────────────

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")
GITHUB_REPO  = os.getenv("GITHUB_REPO", "mshorasoft/Orquesta")
OWNER_EMAIL  = os.getenv("OWNER_EMAIL", "ms.horasoft@gmail.com")

# Almacén de errores y feedbacks recientes
_error_log: list = []
_feedback_log: list = []
_pending_improvements: dict = {}  # id -> propuesta

def log_error(endpoint: str, error: str, context: str = ""):
    """Registra un error para análisis posterior."""
    _error_log.append({
        "ts": datetime.now(timezone.utc).isoformat(),
        "endpoint": endpoint,
        "error": error[:200],
        "context": context[:100]
    })
    # Mantener solo los últimos 100 errores
    if len(_error_log) > 100:
        _error_log.pop(0)

async def send_improvement_email(proposal: dict) -> bool:
    """
    Envía email al dueño con la propuesta de mejora.
    Cascada: Resend → SendGrid → Gmail SMTP → log en consola
    """
    RESEND_KEY    = os.getenv("RESEND_API_KEY", "")
    SENDGRID_KEY  = os.getenv("SENDGRID_API_KEY", "")
    GMAIL_PASS    = os.getenv("GMAIL_APP_PASSWORD", "")   # contraseña de app de Gmail
    
    approve_url = f"{APP_URL}/api/self-improve/approve/{proposal['id']}"
    reject_url  = f"{APP_URL}/api/self-improve/reject/{proposal['id']}"
    
    html_body = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="font-family:sans-serif;background:#f5f5f5;padding:20px;">
<div style="max-width:600px;margin:0 auto;background:white;border-radius:12px;padding:30px;border:1px solid #e0e0e0;">
  <h2 style="color:#1D9E75;margin-top:0;">🤖 Orquesta detectó una mejora</h2>
  <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
    <tr><td style="padding:8px;color:#666;width:120px;">Tipo</td><td style="padding:8px;font-weight:600;">{proposal.get('type','')}</td></tr>
    <tr style="background:#f9f9f9;"><td style="padding:8px;color:#666;">Descripción</td><td style="padding:8px;">{proposal.get('description','')}</td></tr>
    <tr><td style="padding:8px;color:#666;">Impacto</td><td style="padding:8px;">{proposal.get('impact','')}</td></tr>
  </table>
  <div style="background:#1c1e1b;padding:15px;border-radius:8px;margin-bottom:20px;">
    <code style="color:#1D9E75;font-size:13px;white-space:pre-wrap;">{proposal.get('code_summary','')[:500]}</code>
  </div>
  <div style="display:flex;gap:12px;">
    <a href="{approve_url}" style="background:#1D9E75;color:white;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:600;display:inline-block;">✅ Aprobar</a>
    <a href="{reject_url}" style="background:#c0392b;color:white;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:600;display:inline-block;">❌ Rechazar</a>
  </div>
  <p style="color:#999;font-size:12px;margin-top:20px;">Orquesta AI · Auto-mejoramiento · ID: {proposal['id']}</p>
</div>
</body></html>"""

    subject = f"🤖 Orquesta: {proposal.get('description','mejora detectada')[:60]}"
    print(f"📧 Enviando notificación de mejora {proposal['id']} a {OWNER_EMAIL}...")

    # ── 1. Resend ─────────────────────────────────────────────────────────────
    # NOTA: En plan gratuito de Resend, "from" debe ser onboarding@resend.dev
    # y solo funciona si el destinatario está verificado en tu cuenta Resend.
    # Para usar tu propio dominio, verificarlo en resend.com/domains
    RESEND_FROM = os.getenv("RESEND_FROM_EMAIL", "onboarding@resend.dev")
    if RESEND_KEY:
        try:
            async with httpx.AsyncClient(timeout=15) as c:
                r = await c.post(
                    "https://api.resend.com/emails",
                    headers={"Authorization": f"Bearer {RESEND_KEY}", "Content-Type": "application/json"},
                    json={
                        "from": f"Orquesta AI <{RESEND_FROM}>",
                        "to": [OWNER_EMAIL],
                        "subject": subject,
                        "html": html_body,
                    }
                )
                print(f"📧 Resend: {r.status_code} | {r.text[:200]}")
                if r.is_success:
                    print(f"✅ Email enviado via Resend a {OWNER_EMAIL}")
                    return True
                else:
                    print(f"⚠️ Resend falló ({r.status_code}): {r.text[:100]}, probando SendGrid...")
        except Exception as e:
            print(f"⚠️ Resend exception: {e}, probando SendGrid...")

    # ── 2. SendGrid (con sender dinámico — usa el OWNER_EMAIL como from) ───────
    if SENDGRID_KEY:
        # SendGrid permite usar cualquier email si está en "Single Sender Verification"
        # o si el dominio está verificado. Usamos noreply@ como fallback seguro.
        sender_email = os.getenv("SENDGRID_FROM_EMAIL", OWNER_EMAIL)
        try:
            async with httpx.AsyncClient(timeout=15) as c:
                r = await c.post(
                    "https://api.sendgrid.com/v3/mail/send",
                    headers={"Authorization": f"Bearer {SENDGRID_KEY}", "Content-Type": "application/json"},
                    json={
                        "personalizations": [{"to": [{"email": OWNER_EMAIL, "name": "Horacio"}]}],
                        "from": {"email": sender_email, "name": "Orquesta AI"},
                        "reply_to": {"email": sender_email, "name": "Orquesta AI"},
                        "subject": subject,
                        "content": [{"type": "text/html", "value": html_body}],
                        "tracking_settings": {
                            "click_tracking": {"enable": False},
                            "open_tracking": {"enable": False}
                        },
                        "mail_settings": {
                            "bypass_spam_management": {"enable": True}
                        }
                    }
                )
                print(f"📧 SendGrid: {r.status_code} | {r.text[:150]}")
                if r.is_success:
                    print(f"✅ Email enviado via SendGrid a {OWNER_EMAIL}")
                    return True
                else:
                    print(f"⚠️ SendGrid falló: {r.status_code} - {r.text[:200]}")
        except Exception as e:
            print(f"⚠️ SendGrid exception: {e}")

    # ── 3. Gmail SMTP directo (con App Password) ───────────────────────────────
    if GMAIL_PASS:
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"]    = OWNER_EMAIL
            msg["To"]      = OWNER_EMAIL
            msg.attach(MIMEText(html_body, "html"))
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(OWNER_EMAIL, GMAIL_PASS)
                smtp.sendmail(OWNER_EMAIL, OWNER_EMAIL, msg.as_string())
            print(f"✅ Email enviado via Gmail SMTP a {OWNER_EMAIL}")
            return True
        except Exception as e:
            print(f"⚠️ Gmail SMTP exception: {e}")

    # ── 4. Fallback: solo log en consola (siempre funciona) ────────────────────
    print(f"""
╔══════════════════════════════════════════════════════════╗
║  🤖 MEJORA PENDIENTE — {proposal['id']}
║  Tipo: {proposal.get('type','')}
║  Descripción: {proposal.get('description','')[:60]}
║  Aprobar: {approve_url}
║  Rechazar: {reject_url}
╚══════════════════════════════════════════════════════════╝""")
    return False

async def apply_github_change(filename: str, new_content: str, commit_msg: str) -> bool:
    """Aplica un cambio directamente en GitHub via API."""
    if not GITHUB_TOKEN:
        return False
    try:
        # Obtener el SHA actual del archivo
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(
                f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filename}",
                headers={"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
            )
            if not r.is_success:
                return False
            sha = r.json().get("sha", "")
            
            # Actualizar el archivo
            import base64 as _b64
            encoded = _b64.b64encode(new_content.encode()).decode()
            r2 = await c.put(
                f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filename}",
                headers={"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"},
                json={
                    "message": f"🤖 Auto-mejora: {commit_msg}",
                    "content": encoded,
                    "sha": sha,
                    "branch": "main"
                }
            )
            return r2.is_success
    except Exception as e:
        print(f"GitHub apply error: {e}")
        return False

async def get_current_routes_content() -> str:
    """Obtiene el contenido actual de routes.py desde GitHub."""
    if not GITHUB_TOKEN:
        return ""
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.get(
                f"https://api.github.com/repos/{GITHUB_REPO}/contents/app/routes.py",
                headers={"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
            )
            if r.is_success:
                import base64 as _b64
                return _b64.b64decode(r.json()["content"]).decode("utf-8")
    except Exception as e:
        print(f"Error obteniendo routes.py: {e}")
    return ""

async def analyze_and_propose_improvements():
    """
    Analiza errores recientes y propone mejoras.
    Cascada: Gemini → Groq → fallback garantizado.
    Siempre envía email si hay errores para reportar.
    """
    if not _error_log and not _feedback_log:
        print("📋 Sin errores para analizar")
        return None

    top_error = _error_log[-1] if _error_log else _feedback_log[-1] if _feedback_log else {"endpoint": "general", "error": "feedback de usuario"}
    recent = _error_log[-5:]
    errores_txt = "\n".join([f"- {e['endpoint']}: {e['error'][:80]}" for e in recent])

    # Obtener código actual de routes.py para que la IA pueda proponer cambios reales
    current_code_snippet = ""
    if GITHUB_TOKEN:
        try:
            full_code = await get_current_routes_content()
            ep = top_error.get("endpoint", "")
            err_text = (top_error.get("error", "") + " " + top_error.get("context", "")).lower()

            # Buscar fragmento relevante por endpoint Y por palabras clave del error
            search_map = [
                (["image", "vision", "upload", "foto", "imagen", "edit"], "async def call_gemini_vision"),
                (["video", "kling", "fal", "mp4"], "async def generate_video_smart"),
                (["sound", "music", "cancion", "canción", "audio", "udio", "pollination"], "async def generate_music_smart"),
                (["feedback", "auto-mejor", "autorepar", "self-improv", "mejora", "código exacto"], "async def analyze_and_propose_improvements"),
                (["classify", "clasificador", "interpreta", "confunde", "video gen", "sound_gen"], "def classify"),
                (["excel", "xlsx", "word", "docx", "pdf", "archivo"], "async def generate_file_from_prompt"),
                (["upload", "stt", "tts", "voz"], "async def call_openai_tts"),
            ]

            combined = ep + " " + err_text
            for keywords, func_name in search_map:
                if any(k in combined for k in keywords):
                    idx = full_code.find(func_name)
                    if idx >= 0:
                        current_code_snippet = full_code[idx:idx+800]
                        print(f"📎 Fragmento encontrado: {func_name}")
                        break

            # Fallback: primeras líneas del archivo
            if not current_code_snippet and full_code:
                current_code_snippet = full_code[:600]
        except Exception as ce:
            print(f"⚠️ No se pudo obtener código actual: {ce}")

    # Estrategia: nosotros proveemos el old_code exacto, la IA solo propone el new_code
    # Esto garantiza que el match en GitHub siempre funcione
    exact_old_code = current_code_snippet[:600] if current_code_snippet else ""

    code_context = f"""\n\nCÓDIGO ACTUAL EXACTO a mejorar (app/routes.py):\n```python\n{exact_old_code}\n```""" if exact_old_code else ""

    prompt = f"""Sos un experto senior en Python/FastAPI. Hay errores en Orquesta AI que necesitan fix urgente.

Errores recientes:
{errores_txt}{code_context}

TU TAREA:
1. Analizá el error y el código actual
2. Escribí ÚNICAMENTE el código corregido para reemplazar el fragmento mostrado (new_code)
3. El old_code ya está determinado — usá EXACTAMENTE el texto del "CÓDIGO ACTUAL EXACTO" de arriba
4. Si el fix es pequeño y seguro, marcá safe_to_auto_apply como true

IMPORTANTE: new_code debe ser un reemplazo funcional del fragmento mostrado. Mantené el mismo nivel de indentación.

Respondé ÚNICAMENTE con JSON válido (sin markdown, sin texto extra):
{{"type":"bug_fix","description":"descripción clara del problema y fix en español (2-3 oraciones)","impact":"alto","code_summary":"función exacta modificada en app/routes.py","old_code":{json.dumps(exact_old_code)},"new_code":"código corregido aquí","safe_to_auto_apply":false}}"""

    result_text = None

    # 1. Gemini (más quota disponible)
    if GEMINI_KEY:
        try:
            result_text = await asyncio.wait_for(call_gemini(prompt), timeout=15.0)
            print("✅ analyze: Gemini OK")
        except Exception as e:
            print(f"⚠️ analyze Gemini: {e}")

    # 2. Groq fallback
    if not result_text and GROQ_KEY:
        try:
            msgs = [
                {"role": "system", "content": "Respondés SOLO con JSON válido. Sin markdown."},
                {"role": "user", "content": prompt}
            ]
            result_text = await asyncio.wait_for(
                call_groq(msgs, "llama-3.1-8b-instant"), timeout=12.0
            )
            print("✅ analyze: Groq OK")
        except Exception as e:
            print(f"⚠️ analyze Groq: {e}")

    # Parsear JSON
    proposal_data = None
    if result_text:
        try:
            clean = result_text.strip().replace("```json","").replace("```","").strip()
            s, e2 = clean.find("{"), clean.rfind("}") + 1
            if s >= 0 and e2 > s:
                proposal_data = json.loads(clean[s:e2])
        except Exception as e:
            print(f"⚠️ analyze JSON parse: {e}")

    # Construir propuesta — con datos de IA o fallback descriptivo
    proposal_id = str(uuid.uuid4())[:8]
    if proposal_data:
        proposal = {"id": proposal_id, "ts": datetime.now(timezone.utc).isoformat(),
                    "status": "pending", **proposal_data}
    else:
        # Fallback garantizado — siempre hay algo para enviar
        proposal = {
            "id": proposal_id,
            "ts": datetime.now(timezone.utc).isoformat(),
            "status": "pending",
            "type": "revision_manual",
            "description": f"Error en {top_error['endpoint']}: {top_error['error'][:100]}",
            "impact": "requiere revisión",
            "code_summary": f"Revisar {top_error['endpoint']} — {top_error.get('context','sin contexto')}",
            "old_code": "",
            "new_code": "",
            "safe_to_auto_apply": False,
        }

    _pending_improvements[proposal_id] = proposal
    print(f"✅ Mejora propuesta: {proposal_id} — {proposal.get('description','')[:60]}")

    # Enviar email siempre
    asyncio.create_task(send_improvement_email(proposal))
    return proposal_id


# ── ENDPOINTS DEL SISTEMA DE AUTO-MEJORAMIENTO ────────────────────────────────

@router.get("/self-improve/pending")
async def get_pending_improvements():
    """Lista las mejoras pendientes de aprobación."""
    return {
        "pending": list(_pending_improvements.values()),
        "recent_errors": len(_error_log),
        "total_proposed": len(_pending_improvements)
    }

@router.get("/self-improve/approve/{proposal_id}")
async def approve_improvement(proposal_id: str):
    """Aprueba y aplica automáticamente una mejora via GitHub."""
    from fastapi.responses import HTMLResponse

    if proposal_id not in _pending_improvements:
        return HTMLResponse("<html><body style='font-family:sans-serif;background:#0d0f0d;color:#e3e8e4;padding:2rem;'><h2 style='color:#c0392b;'>❌ Propuesta no encontrada o ya procesada</h2></body></html>")

    proposal = _pending_improvements[proposal_id]
    proposal["status"] = "approved"
    proposal["approved_at"] = datetime.now(timezone.utc).isoformat()
    print(f"✅ MEJORA APROBADA: {proposal_id} - {proposal.get('description','')}")

    applied = False
    error_msg = ""

    # Aplicar el cambio en GitHub si hay old_code y new_code
    old_code = proposal.get("old_code", "")
    new_code = proposal.get("new_code", "")

    if old_code and new_code and GITHUB_TOKEN:
        try:
            # Obtener el contenido actual del archivo
            current_content = await get_current_routes_content()
            if current_content and old_code.strip() in current_content:
                # Aplicar el reemplazo exacto
                updated_content = current_content.replace(old_code.strip(), new_code.strip(), 1)
                commit_msg = f"Auto-mejora #{proposal_id}: {proposal.get('description','mejora automática')}"
                applied = await apply_github_change("app/routes.py", updated_content, commit_msg)
                if applied:
                    print(f"✅ Cambio aplicado en GitHub: {proposal_id}")
                else:
                    error_msg = "Error al hacer commit en GitHub"
            elif current_content:
                # Fragmento no encontrado exacto — agregar mejora como comentario documentado
                improvement_note = f"""
# ── AUTO-MEJORA #{proposal_id} (pendiente de aplicar manualmente) ──────────
# Descripción: {proposal.get('description','')}
# Fecha: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}
# Cambio sugerido:
# ANTES: {old_code.strip()[:200]}
# DESPUÉS: {new_code.strip()[:200]}
# ─────────────────────────────────────────────────────────────────────────────
"""
                # Agregar al final del archivo como referencia
                updated_content = current_content + improvement_note
                commit_msg = f"Auto-mejora #{proposal_id} (doc): {proposal.get('description','')}"
                applied = await apply_github_change("app/routes.py", updated_content, commit_msg)
                if applied:
                    print(f"✅ Mejora documentada en GitHub: {proposal_id}")
                    error_msg = "Mejora documentada — fragmento exacto no encontrado. Requiere revisión manual."
                else:
                    error_msg = "No se pudo hacer commit en GitHub"
            else:
                error_msg = "No se pudo obtener el código actual de GitHub"
                print(f"⚠️ {error_msg}")
        except Exception as e:
            error_msg = str(e)[:100]
            print(f"❌ Error aplicando cambio: {e}")

    ts_fmt = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")

    if applied:
        status_color = "#1D9E75"; status_icon = "✅"; status_title = "¡Mejora aplicada!"
        status_msg = "El cambio fue commiteado en GitHub. Railway lo detectará y redesplegará automáticamente en ~2 minutos."
        deploy_badge = """<div style="background:#0a2e1e;border:1px solid #1D9E75;border-radius:10px;padding:1rem 1.5rem;margin:1.5rem 0;text-align:left;">
          <div style="color:#1D9E75;font-size:13px;font-weight:700;margin-bottom:6px;">🚀 Deploy automático en progreso</div>
          <div style="color:#7abf9e;font-size:12px;">GitHub → Railway detecta el commit → redeploy en ~2 min</div>
          <div style="margin-top:10px;height:4px;background:#1a3a2a;border-radius:4px;overflow:hidden;">
            <div style="height:100%;width:100%;background:linear-gradient(90deg,#1D9E75,#2dd4a0);animation:bar 2s ease-in-out infinite;"></div>
          </div></div>
        <style>@keyframes bar{{0%{{width:0%}}100%{{width:100%}}}}</style>"""
    else:
        status_color = "#f39c12"; status_icon = "⚠️"; status_title = "Mejora registrada"
        status_msg = error_msg if error_msg else "La mejora fue aprobada pero requiere aplicación manual (no había código exacto para reemplazar automáticamente)."
        deploy_badge = f"""<div style="background:#2a1e0a;border:1px solid #f39c12;border-radius:10px;padding:1rem 1.5rem;margin:1.5rem 0;text-align:left;">
          <div style="color:#f39c12;font-size:13px;font-weight:700;margin-bottom:6px;">📋 Acción manual requerida</div>
          <div style="color:#c9a85c;font-size:12px;">{status_msg}</div></div>"""

    desc = proposal.get("description","Sin descripción")
    code_sum = proposal.get("code_summary","")
    impact = proposal.get("impact","medio")
    impact_color = "#e74c3c" if impact=="alto" else "#f39c12" if impact=="medio" else "#2ecc71"

    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Orquesta · Mejora aprobada</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d0f0d;color:#e3e8e4;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:1.5rem}}
.card{{background:#141614;border:1px solid #2a302a;border-radius:16px;padding:2.5rem 2rem;max-width:560px;width:100%}}
.logo{{color:#1D9E75;font-size:13px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;margin-bottom:1.5rem;opacity:.7}}
.icon{{font-size:3rem;margin-bottom:.75rem}}.badge{{display:inline-block;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700;background:{impact_color}22;color:{impact_color};border:1px solid {impact_color}44;margin-bottom:1.2rem}}
h1{{font-size:1.6rem;color:{status_color};margin-bottom:.5rem}}.desc{{color:#b0c0b2;font-size:14px;line-height:1.6;margin-bottom:.75rem}}
.code-sum{{background:#0d1a0f;border-left:3px solid #1D9E75;padding:.75rem 1rem;border-radius:0 8px 8px 0;font-size:12px;color:#7abf9e;font-family:monospace;margin-bottom:1rem}}
.meta{{color:#4a5e4c;font-size:11px;margin-top:1.5rem}}.btn{{display:inline-block;margin-top:1.75rem;padding:10px 24px;background:#1D9E75;color:white;border-radius:8px;text-decoration:none;font-size:13px;font-weight:600}}</style>
</head><body><div class="card">
<div class="logo">🤖 Orquesta · Auto-mejoramiento</div>
<div class="icon">{status_icon}</div><h1>{status_title}</h1>
<span class="badge">Impacto: {impact}</span>
<p class="desc">{desc}</p>
{"<div class='code-sum'>" + code_sum + "</div>" if code_sum else ""}
{deploy_badge}
<div class="meta">ID: {proposal_id} &nbsp;·&nbsp; {ts_fmt}</div>
<a class="btn" href="/api/self-improve/pending">Ver todas las mejoras →</a>
</div></body></html>""")

@router.get("/self-improve/reject/{proposal_id}")
async def reject_improvement(proposal_id: str):
    """Rechaza una mejora propuesta."""
    if proposal_id not in _pending_improvements:
        return {"error": "Propuesta no encontrada"}
    
    _pending_improvements[proposal_id]["status"] = "rejected"
    print(f"❌ MEJORA RECHAZADA: {proposal_id}")
    
    from fastapi.responses import HTMLResponse
    desc_r = _pending_improvements.get(proposal_id, {}).get("description", "Sin descripción")
    ts_r = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Orquesta · Mejora rechazada</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d0f0d;color:#e3e8e4;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:1.5rem}}
.card{{background:#141614;border:1px solid #2a302a;border-radius:16px;padding:2.5rem 2rem;max-width:480px;width:100%;text-align:center}}
.logo{{color:#1D9E75;font-size:13px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;margin-bottom:1.5rem;opacity:.7}}
.icon{{font-size:3rem;margin-bottom:.75rem}}h1{{font-size:1.6rem;color:#c0392b;margin-bottom:.75rem}}
.desc{{color:#7a8a7c;font-size:13px;line-height:1.6;margin-bottom:1rem}}
.note{{background:#1a0d0d;border:1px solid #c0392b33;border-radius:10px;padding:.75rem 1rem;font-size:12px;color:#a07070;margin-bottom:1.5rem}}
.meta{{color:#3a4e3c;font-size:11px;margin-top:1rem}}.btn{{display:inline-block;margin-top:1.5rem;padding:10px 24px;background:#1D9E75;color:white;border-radius:8px;text-decoration:none;font-size:13px;font-weight:600}}</style>
</head><body><div class="card">
<div class="logo">🤖 Orquesta · Auto-mejoramiento</div>
<div class="icon">❌</div><h1>Mejora rechazada</h1>
<p class="desc">{desc_r}</p>
<div class="note">Orquesta tomó nota. Este cambio no se aplicará.<br>Si fue un error, podés re-disparar el análisis manualmente.</div>
<div class="meta">ID: {proposal_id} &nbsp;·&nbsp; {ts_r}</div>
<a class="btn" href="/api/self-improve/pending">Ver todas las mejoras →</a>
</div></body></html>""")

@router.post("/self-improve/trigger")
async def trigger_analysis_post():
    """Dispara un análisis manual de mejoras via POST."""
    # Agregar algunos errores de ejemplo si no hay ninguno para probar
    if not _error_log:
        log_error("/api/orchestrate", "JWT verify failed — Token inválido", "auth")
        log_error("/api/stt", "Quota exceeded — OpenAI sin créditos", "stt")
        log_error("/api/orchestrate", "Gemini 503 UNAVAILABLE — model saturado", "vision")
    await analyze_and_propose_improvements()
    return {"status": "analysis_triggered", "errors_analyzed": len(_error_log), "pending": len(_pending_improvements)}

@router.get("/self-improve/trigger")
async def trigger_analysis_get():
    """Dispara un análisis manual via GET — responde rápido."""
    from fastapi.responses import HTMLResponse

    # Agregar errores de ejemplo si no hay ninguno
    if not _error_log:
        log_error("/api/upload", "Gemini 429 quota exceeded", "vision")
        log_error("/api/stt", "OpenAI quota exceeded", "stt")
        log_error("/api/orchestrate", "video_gen no credits", "video")

    # Analizar errores — usar Gemini primero (más tokens disponibles), Groq como fallback
    top_err = _error_log[-1] if _error_log else {"endpoint": "api/general", "error": "mejora de rendimiento general"}

    # Usar la misma función que el análisis automático — incluye descarga de código desde GitHub
    proposal_id = await analyze_and_propose_improvements()
    proposal = _pending_improvements.get(proposal_id) if proposal_id else None
    email_sent = bool(proposal)
    print(f"📧 Trigger GET: {'✅ análisis completado' if proposal_id else '❌ sin errores para analizar'}")

    pending = list(_pending_improvements.values())
    items = "".join([
        f"""<li style='margin:.8rem 0;padding:.8rem;background:#1c1e1b;border-radius:8px;border-left:3px solid #1D9E75;'>
        <strong style='color:#1D9E75;'>{p.get('type','').upper()}</strong>: {p.get('description','')}
        <br><small style='color:#aaa;'>{p.get('code_summary','')}</small>
        <br style='margin:.4rem 0;'>
        <a href='/api/self-improve/approve/{p["id"]}' style='color:#1D9E75;margin-right:1rem;font-weight:600;'>✅ Aprobar</a>
        <a href='/api/self-improve/reject/{p["id"]}' style='color:#c0392b;font-weight:600;'>❌ Rechazar</a>
        </li>"""
        for p in pending
    ])

    return HTMLResponse(f"""<!DOCTYPE html>
    <html><head><meta charset="utf-8">
    <style>body{{font-family:system-ui,sans-serif;background:#0d0f0d;color:#e3e8e4;padding:2rem;max-width:700px;margin:0 auto;}}
    h2{{color:#1D9E75;}} a{{color:#1D9E75;}} .stat{{background:#1c1e1b;padding:.5rem 1rem;border-radius:6px;display:inline-block;margin:.3rem;}}</style>
    <script>setTimeout(()=>location.reload(),8000);</script>
    </head><body>
    <h2>🤖 Sistema de Auto-mejoramiento — Orquesta AI</h2>
    <div>
      <span class="stat">📊 Errores registrados: <strong>{len(_error_log)}</strong></span>
      <span class="stat">💡 Mejoras pendientes: <strong>{len(pending)}</strong></span>
    </div>
    <p style="color:#aaa;font-size:13px;margin-top:1rem;">
      📧 Email {'✅ enviado' if email_sent else '❌ no enviado — revisar logs de Railway'} a {OWNER_EMAIL}<br>
      🔄 La página se actualiza sola en 8 segundos.
    </p>
    {"<h3 style='color:#1D9E75;margin-top:1.5rem;'>💡 Mejoras detectadas</h3><ul style='padding:0;list-style:none;'>" + items + "</ul>" if pending else ""}
    <br><a href="/api/self-improve/pending">Ver JSON →</a> · 
    <a href="/">Volver a Orquesta</a>
    </body></html>""")

@router.post("/orchestrate/stream")
async def orchestrate_stream(req: OrchestrateReq, request: Request, authorization: str = Header(None)):
    """Versión streaming del orchestrate — devuelve tokens en tiempo real."""
    from fastapi.responses import StreamingResponse as SR
    
    if not req.prompt.strip():
        raise HTTPException(400, "Prompt vacío")

    # Rate limiting
    client_ip = request.client.host if request.client else "unknown"
    if not check_rate_limit(client_ip, max_per_minute=30):
        raise HTTPException(429, "Demasiadas peticiones. Esperá un momento.")

    user = await get_optional_user(authorization)
    if not user and supabase:
        for field, value in [("auth_id", req.auth_id), ("auth_id", req.user_id), ("id", req.user_id)]:
            if not value: continue
            try:
                result = supabase.table("users").select("*").eq(field, value).single().execute()
                if result.data: user = result.data; break
            except: continue

    task = classify(req.prompt, req.mode, req.history)
    block = check_pro_access(user, task)
    if block and block.get("blocked"):
        async def blocked_gen():
            yield f"data: {json.dumps({'text': block['message'], 'done': True})}\n\n"
        return SR(blocked_gen(), media_type="text/event-stream")

    username = user["name"] if user else req.username
    system = get_system(req.mode, username)

    # ── Tareas especiales: video, imagen, archivos — no son streaming de texto ──
    if task == "video_gen":
        async def video_stream_gen():
            yield f"data: {json.dumps({'text': '⏳ Generando tu video con IA...', 'done': False})}\n\n"
            try:
                video_url, result, label = await generate_video_smart(
                    req.prompt, req.history, req.mode, username,
                    user=user,
                    model_key=req.video_model,
                    duration_s=req.video_duration,
                )
                payload = {"text": result, "done": True, "label": label}
                if video_url:
                    payload["video_url"] = video_url
                yield f"data: {json.dumps(payload)}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'text': f'Error generando video: {str(e)}', 'done': True})}\n\n"
        return SR(video_stream_gen(), media_type="text/event-stream",
                  headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    if task == "image_gen":
        async def image_stream_gen():
            yield f"data: {json.dumps({'text': '🎨 Generando imagen...', 'done': False})}\n\n"
            try:
                img_url, result, label = await generate_image_smart(req.prompt)
                payload = {"text": result, "done": True, "label": label}
                if img_url:
                    payload["image_url"] = img_url
                yield f"data: {json.dumps(payload)}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'text': f'Error generando imagen: {str(e)}', 'done': True})}\n\n"
        return SR(image_stream_gen(), media_type="text/event-stream",
                  headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    # ── Streaming de texto normal (general, code, realtime, etc.) ─────────────
    model = TASK_MODELS.get(task, "llama-3.3-70b-versatile")
    msgs = build_messages(system, req.history, req.prompt)

    # Si es realtime, enriquecer con búsqueda web antes de streamear
    if task == "realtime" and TAVILY_KEY:
        try:
            ctx = await deep_search(req.prompt)
            if ctx:
                enriched = (f'El usuario pregunta: "{req.prompt}"\n\n'
                            f"=== INFORMACIÓN DE MÚLTIPLES FUENTES ===\n{ctx}\n\n"
                            f"Respondé de forma COMPLETA Y DETALLADA integrando todos los datos.")
                msgs = build_messages(system, req.history[:-1] if req.history else [], enriched)
        except Exception:
            pass

    async def stream_gen():
        try:
            async with httpx.AsyncClient(timeout=60) as c:
                async with c.stream("POST", "https://api.groq.com/openai/v1/chat/completions",
                    headers={"Authorization": f"Bearer {GROQ_KEY}"},
                    json={"model": model, "messages": msgs, "max_tokens": 4096,
                          "temperature": 0.7, "stream": True}) as resp:
                    async for line in resp.aiter_lines():
                        if line.startswith("data: "):
                            data = line[6:]
                            if data == "[DONE]":
                                yield f"data: {json.dumps({'text': '', 'done': True})}\n\n"
                                break
                            try:
                                chunk = json.loads(data)
                                token = chunk["choices"][0]["delta"].get("content", "")
                                if token:
                                    yield f"data: {json.dumps({'text': token, 'done': False})}\n\n"
                            except: pass
        except Exception as e:
            yield f"data: {json.dumps({'text': f'Error: {str(e)}', 'done': True})}\n\n"

    return SR(stream_gen(), media_type="text/event-stream",
              headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})




# ─────────────────────────────────────────────────────────────────────────────
#  ENDPOINTS DE CRÉDITOS DE VIDEO
# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
#  RECARGA DE SALDO DE VIDEO — El usuario paga, Orquesta acredita USD
# ─────────────────────────────────────────────────────────────────────────────

# Packs predefinidos (precio al usuario, saldo que se acredita)
VIDEO_PACKS = {
    "pack_5":  {"usd": 5.00,  "label": "Pack $5  — ~7 videos de 5s"},
    "pack_10": {"usd": 10.00, "label": "Pack $10 — ~14 videos de 5s"},
    "pack_20": {"usd": 20.00, "label": "Pack $20 — ~28 videos de 5s"},
    "pack_50": {"usd": 50.00, "label": "Pack $50 — ~70 videos de 5s"},
}

class VideoTopupReq(BaseModel):
    pack_id: str       # "pack_5" | "pack_10" | "pack_20" | "pack_50"
    user_id: str = ""
    user_email: str = ""
    access_token: str = ""

@router.post("/video/topup")
async def video_topup(req: VideoTopupReq, authorization: str = Header(None)):
    """Crea preferencia de MercadoPago para recargar saldo de video."""
    if not MP_ACCESS_TOKEN:
        raise HTTPException(500, "MercadoPago no configurado")

    pack = VIDEO_PACKS.get(req.pack_id)
    if not pack:
        raise HTTPException(400, f"Pack inválido: {req.pack_id}")

    # Obtener usuario
    user = await get_optional_user(authorization)
    if not user and req.user_id and supabase:
        try:
            result = supabase.table("users").select("*").eq("id", req.user_id).single().execute()
            if result.data: user = result.data
        except: pass
    if not user and req.user_email:
        user = {"id": req.user_id or str(uuid.uuid4()), "email": req.user_email, "name": "", "auth_id": ""}
    if not user:
        raise HTTPException(401, "No autenticado")

    # Crear preferencia en MercadoPago
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post(
            "https://api.mercadopago.com/checkout/preferences",
            headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json={
                "items": [{
                    "title": f"Orquesta AI — {pack['label']}",
                    "quantity": 1,
                    "unit_price": pack["usd"],
                    "currency_id": "USD",
                }],
                "payer": {"email": user.get("email", "")},
                "external_reference": f"video_topup|{user['id']}|{req.pack_id}|{pack['usd']}",
                "metadata": {
                    "user_id": user["id"],
                    "pack_id": req.pack_id,
                    "amount_usd": pack["usd"],
                    "type": "video_topup",
                },
                "back_urls": {
                    "success": f"{APP_URL}/?video_topup=success&pack={req.pack_id}",
                    "failure": f"{APP_URL}/?video_topup=failed",
                    "pending": f"{APP_URL}/?video_topup=pending",
                },
                "auto_return": "approved",
                "notification_url": f"{APP_URL}/api/webhooks/mercadopago",
            }
        )
        d = r.json()
        if not r.is_success:
            raise HTTPException(502, f"Error MercadoPago: {d.get('message', 'Unknown')}")
        return {
            "checkout_url": d.get("init_point") or d.get("sandbox_init_point"),
            "preference_id": d.get("id"),
            "pack": pack,
        }



# ─────────────────────────────────────────────────────────────────────────────
#  VERIFICACIÓN MANUAL DE PAGO MP — el usuario lo llama al volver de MP
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/video/verify-payment")
async def verify_video_payment(data: dict, authorization: str = Header(None)):
    """
    El frontend lo llama cuando el usuario vuelve de MercadoPago.
    Verifica el pago directo en MP API y acredita si está aprobado.
    """
    user = await get_optional_user(authorization)
    if not user and SUPABASE_URL and SUPABASE_SERVICE_KEY:
        try:
            token = (authorization or "").replace("Bearer ", "").strip()
            parts = token.split(".")
            if len(parts) == 3:
                pad = parts[1] + "=" * (4 - len(parts[1]) % 4)
                jwt_p = json.loads(base64.urlsafe_b64decode(pad))
                auth_id = jwt_p.get("sub", "")
                if auth_id:
                    async with httpx.AsyncClient(timeout=8) as hx:
                        r = await hx.get(
                            f"{SUPABASE_URL}/rest/v1/users?auth_id=eq.{auth_id}&select=*&limit=1",
                            headers={"apikey": SUPABASE_SERVICE_KEY, "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}"}
                        )
                        if r.is_success and r.json():
                            user = r.json()[0]
        except Exception as e:
            print(f"verify-payment auth fallback: {e}")

    if not user:
        raise HTTPException(401, "No autenticado")

    pack_id = data.get("pack_id", "")
    preference_id = data.get("preference_id", "")

    if not MP_ACCESS_TOKEN:
        raise HTTPException(500, "MercadoPago no configurado")

    # Buscar pagos recientes del usuario en MP
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            # Buscar por external_reference que contiene user_id
            search_url = f"https://api.mercadopago.com/v1/payments/search"
            r = await c.get(search_url, headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"},
                params={"external_reference": f"video_topup|{user['id']}", "sort": "date_created", "criteria": "desc", "range": "date_created", "begin_date": "NOW-1DAYS", "end_date": "NOW"})
            results = r.json().get("results", []) if r.is_success else []

            # Buscar el pago aprobado más reciente
            approved = [p for p in results if p.get("status") == "approved"]
            if not approved:
                # También buscar por preference_id si lo tenemos
                current_balance = await get_video_credits(user)
                return {
                    "credited": False,
                    "balance_usd": current_balance["balance_usd"],
                    "message": "Pago pendiente o no encontrado. El webhook llegará en breve."
                }

            payment = approved[0]
            payment_id = str(payment["id"])
            amount_usd = float(payment.get("transaction_amount", 0))

            # Verificar si este pago ya fue acreditado (buscar en payment_events)
            already_credited = False
            try:
                ev = supabase.table("payment_events").select("id").eq("event_id", payment_id).execute()
                already_credited = bool(ev.data)
            except Exception:
                pass

            if not already_credited and amount_usd > 0:
                # Acreditar el saldo
                try:
                    existing = supabase.table("video_balance").select("balance_usd").eq("user_id", user["id"]).execute()
                    if existing.data:
                        new_bal = round(float(existing.data[0].get("balance_usd", 0)) + amount_usd, 4)
                        supabase.table("video_balance").update({"balance_usd": new_bal, "updated_at": datetime.now(timezone.utc).isoformat()}).eq("user_id", user["id"]).execute()
                    else:
                        supabase.table("video_balance").insert({"user_id": user["id"], "balance_usd": amount_usd, "updated_at": datetime.now(timezone.utc).isoformat()}).execute()

                    # Registrar para no duplicar
                    try:
                        supabase.table("payment_events").insert({
                            "provider": "mercadopago",
                            "event_type": "video_topup.verified",
                            "event_id": payment_id,
                            "amount": int(amount_usd * 100),
                            "currency": "USD",
                            "plan": pack_id or "video_topup",
                            "status": "success",
                            "raw_payload": payment,
                        }).execute()
                    except Exception:
                        pass

                    print(f"✅ Pago verificado y acreditado: ${amount_usd} USD → {user['id'][:8]}")
                    credits = await get_video_credits(user)
                    return {"credited": True, "balance_usd": credits["balance_usd"], "amount_credited": amount_usd}

                except Exception as e:
                    print(f"Error acreditando en verify-payment: {e}")
                    raise HTTPException(500, f"Error al acreditar: {str(e)}")

            # Ya estaba acreditado
            credits = await get_video_credits(user)
            return {"credited": already_credited, "balance_usd": credits["balance_usd"], "message": "Ya acreditado previamente"}

    except HTTPException:
        raise
    except Exception as e:
        print(f"verify-payment error: {e}")
        raise HTTPException(502, f"Error verificando pago: {str(e)[:100]}")

@router.get("/video/balance")
async def get_video_balance(authorization: str = Header(None)):
    """Retorna el saldo de video del usuario en USD. Nunca lanza error."""
    prices_table = [
        {
            "model_key": mk,
            "duration_s": dur,
            "label": info["label"],
            "user_price_usd": info["user_price"],
            "can_afford": False,  # se actualiza abajo
        }
        for (mk, dur), info in VIDEO_PRICES.items()
    ]
    prices_table.sort(key=lambda x: x["user_price_usd"])

    try:
        user = await get_optional_user(authorization)

        # Fallback JWT si get_optional_user falló
        if not user and authorization and authorization.startswith("Bearer ") and supabase:
            try:
                token = authorization.replace("Bearer ", "").strip()
                parts = token.split(".")
                if len(parts) == 3:
                    pad = parts[1] + "=" * (4 - len(parts[1]) % 4)
                    jwt_p = json.loads(base64.urlsafe_b64decode(pad))
                    auth_id = jwt_p.get("sub", "")
                    if auth_id:
                        async with httpx.AsyncClient(timeout=8) as hx:
                            r = await hx.get(
                                f"{SUPABASE_URL}/rest/v1/users?auth_id=eq.{auth_id}&select=*&limit=1",
                                headers={"apikey": SUPABASE_SERVICE_KEY, "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}"}
                            )
                            if r.is_success and r.json():
                                user = r.json()[0]
            except Exception as je:
                print(f"video/balance JWT fallback: {je}")

        if not user:
            return {"balance_usd": 0.0, "is_pro": False, "prices": prices_table}

        credits = await get_video_credits(user)
        balance = credits.get("balance_usd", 0.0)
        is_pro = credits.get("is_pro", False)

        # Marcar cuáles puede pagar
        for p in prices_table:
            p["can_afford"] = balance >= p["user_price_usd"]

        return {"balance_usd": balance, "is_pro": is_pro, "prices": prices_table}

    except Exception as e:
        print(f"video/balance endpoint error: {e}")
        # Nunca devolver 500 — el frontend muestra error rojo si falla
        return {"balance_usd": 0.0, "is_pro": False, "prices": prices_table}

@router.get("/video/credits")
async def get_my_video_credits(authorization: str = Header(None)):
    """Retorna créditos de video disponibles para el usuario."""
    user = await get_optional_user(authorization)
    if not user:
        return {"available": 0, "used_this_month": 0, "monthly_allowance": 0, "is_pro": False}
    credits = await get_video_credits(user)
    credits["is_pro"] = user.get("plan") == "pro"
    return credits


@router.post("/video/confirm")
async def confirm_video_generation(data: dict, authorization: str = Header(None)):
    """
    Verifica si el usuario puede generar un video y retorna info de costo.
    El frontend llama esto ANTES de generar, para mostrar el modal de confirmación.
    """
    user = await get_optional_user(authorization)

    # Fallback: buscar por JWT raw si get_optional_user falló
    if not user and authorization and authorization.startswith("Bearer ") and supabase:
        try:
            token = authorization.replace("Bearer ", "").strip()
            parts = token.split(".")
            if len(parts) == 3:
                pad = parts[1] + "=" * (4 - len(parts[1]) % 4)
                jwt_payload = json.loads(base64.urlsafe_b64decode(pad))
                auth_id = jwt_payload.get("sub", "")
                if auth_id:
                    async with httpx.AsyncClient(timeout=10) as hx:
                        url = f"{SUPABASE_URL}/rest/v1/users?auth_id=eq.{auth_id}&select=*&limit=1"
                        r = await hx.get(url, headers={
                            "apikey": SUPABASE_SERVICE_KEY,
                            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}"
                        })
                        if r.is_success and r.json():
                            user = r.json()[0]
                            print(f"✅ /video/confirm: user encontrado via fallback JWT: {user.get('plan')}")
        except Exception as e:
            print(f"/video/confirm JWT fallback error: {e}")

    if not user:
        return {
            "can_generate": False,
            "reason": "no_auth",
            "message": "Iniciá sesión para generar videos.",
        }

    is_pro = (
        user.get("plan") == "pro" and (
            user.get("plan_expires_at") is None or
            parse_expiry(user["plan_expires_at"]) > datetime.now(timezone.utc)
        )
    )
    if not is_pro:
        return {
            "can_generate": False,
            "reason": "not_pro",
            "message": "La generación de videos es exclusiva del Plan Pro.",
        }

    credits = await get_video_credits(user)
    balance = credits["balance_usd"]

    # Construir tabla de precios con saldo disponible
    prices_table = []
    for (mk, dur), info in VIDEO_PRICES.items():
        can_afford = balance >= info["user_price"]
        prices_table.append({
            "model_key": mk,
            "duration_s": dur,
            "label": info["label"],
            "user_price_usd": info["user_price"],
            "fal_cost_usd": info["fal_cost"],
            "can_afford": can_afford,
        })

    # Ordenar por precio
    prices_table.sort(key=lambda x: x["user_price_usd"])

    cheapest_available = next((p for p in prices_table if p["can_afford"]), None)
    can_generate = cheapest_available is not None

    return {
        "can_generate": can_generate,
        "balance_usd": balance,
        "prices": prices_table,
        "reason": "ok" if can_generate else "insufficient_balance",
        "message": (
            f"Saldo disponible: **${balance:.2f} USD**"
            if can_generate
            else f"Saldo insuficiente (${balance:.2f} USD). Cargá créditos para generar videos."
        ),
    }

# ─────────────────────────────────────────────────────────────────────────────
#  VIDEO DOWNLOAD PROXY — Descarga forzada sin CORS issues
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/video/download")
async def download_video(url: str, filename: str = "orquesta_video.mp4"):
    """
    Proxy de descarga para videos generados por Replicate/FAL/ModelsLab.
    Fuerza el header Content-Disposition para que el browser descargue en vez de abrir.
    """
    if not url:
        raise HTTPException(400, "URL requerida")

    # Solo permitir dominios conocidos de generación de video
    allowed_domains = [
        "replicate.delivery", "pbxt.replicate.delivery", "api.replicate.com",
        "fal.media", "cdn.fal.ai", "v3.fal.media",
        "modelslab.com", "cdn.modelslab.com",
        "up.railway.app",  # motor propio
    ]
    from urllib.parse import urlparse
    parsed = urlparse(url)
    if not any(d in (parsed.netloc or "") for d in allowed_domains):
        raise HTTPException(403, f"Dominio no permitido: {parsed.netloc}")

    try:
        async with httpx.AsyncClient(timeout=120, follow_redirects=True) as c:
            r = await c.get(url)
            if not r.is_success:
                raise HTTPException(502, f"Error al obtener el video: HTTP {r.status_code}")

            content_type = r.headers.get("content-type", "video/mp4")
            # Forzar extensión correcta
            if "webm" in content_type:
                filename = filename.replace(".mp4", ".webm")
            elif "quicktime" in content_type or "mov" in content_type:
                filename = filename.replace(".mp4", ".mov")

            return Response(
                content=r.content,
                media_type=content_type,
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Content-Length": str(len(r.content)),
                    "Cache-Control": "no-cache",
                }
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error en proxy de descarga: {str(e)}")

@router.get("/status")
async def status():
    return {
        "groq":        bool(GROQ_KEY),
        "tavily":      bool(TAVILY_KEY),
        "gemini":      bool(GEMINI_KEY),
        "openai":      bool(OPENAI_KEY),
        "supabase":    bool(SUPABASE_URL),
        "stripe":      bool(STRIPE_SECRET_KEY),
        "mercadopago": bool(MP_ACCESS_TOKEN),
        "videogen":    bool(os.getenv("VIDEOGEN_URL","")),
        "file_generation": True,
        "tts": bool(OPENAI_KEY)
    }
